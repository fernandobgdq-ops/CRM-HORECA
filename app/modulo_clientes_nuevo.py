"""
MÓDULO CLIENTES - VISTA 360 PROFESIONAL V2
Sistema CRM con interfaz moderna y KPIs de impacto
MEJORAS:
- Eliminación de interacciones en Timeline
- Servicios contratados reales desde CLIENTES_ACTIVOS
- Dashboard con datos reales (Salud basada en Satisfacción, LTV, Servicio Activo)
- Programar Tarea funcional que va a Próximas Acciones
- Directorio de Contactos completo y funcional
"""

import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import time
import config
import utils
from funciones_leads import crear_nuevo_lead_modal, convertir_lead_a_cliente_modal

# ============================================================================
# FUNCIÓN PRINCIPAL DEL MÓDULO
# ============================================================================

def modulo_clientes():
    """Módulo de gestión de clientes - Vista 360 Profesional"""
    
    # Control de pestaña activa mediante session_state
    if 'tab_activa' not in st.session_state:
        st.session_state.tab_activa = 0
    
    # Tabs principales
    tab1, tab2 = st.tabs([
        "📋 Gestión de Leads",
        "👤 Clientes Actuales"
    ])
    
    with tab1:
        mostrar_lista_clientes()
    
    with tab2:
        mostrar_vista_360_cliente()
    
    # Si hay que cambiar de tab automáticamente, hacerlo con callback
    if st.session_state.get('tab_activa', 0) == 1 and st.session_state.get('programar_tarea', False) == False:
        if st.session_state.get('tarea_programada', False):
            st.session_state['tarea_programada'] = False

# ============================================================================
# TAB 1: LISTA DE CLIENTES
# ============================================================================

def mostrar_lista_clientes():
    """Lista de LEADS con filtros y gestión"""
    st.subheader("📋 Gestión de LEADS")
    
    # Leer LEADS directamente sin caché
    from openpyxl import load_workbook
    try:
        wb = load_workbook(config.ARCHIVO_CRM, data_only=True)
        ws = wb['LEADS']
        
        datos = []
        encabezados = None
        
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                encabezados = row
            else:
                datos.append(row)
        
        if encabezados:
            df_leads = pd.DataFrame(datos, columns=encabezados)
        else:
            df_leads = pd.DataFrame()
        
        wb.close()
    except Exception as e:
        st.error(f"Error al leer LEADS: {str(e)}")
        df_leads = pd.DataFrame()
    
    if df_leads.empty:
        st.info("📝 No hay leads registrados. Crea el primero.")
        if st.button("➕ Nuevo Lead", use_container_width=True, type="primary"):
            st.session_state['crear_nuevo_lead'] = True
            st.rerun()
        return
    
    # Asegurar columnas
    if 'Estado Lead' not in df_leads.columns:
        df_leads['Estado Lead'] = 'Lead'
    if 'Comercial Asignado' not in df_leads.columns:
        df_leads['Comercial Asignado'] = ''
    
    # ========== FILTROS ==========
    col1, col2, col3, col4 = st.columns([2, 2, 2, 1])
    
    with col1:
        filtro_estado = st.multiselect(
            "Estado:",
            options=['Lead', 'Cliente', 'Cliente Activo', 'Pausado', 'Baja'],
            default=[],
            key="filtro_estado_leads"
        )
    
    with col2:
        comerciales = df_leads['Comercial Asignado'].unique()
        comerciales = [c for c in comerciales if pd.notna(c) and c != '']
        filtro_comercial = st.multiselect(
            "Comercial:",
            options=comerciales,
            default=[],
            key="filtro_comercial_leads"
        )
    
    with col3:
        buscar = st.text_input("🔍 Buscar por nombre:", key="buscar_lead")
    
    with col4:
        st.write("")
        st.write("")
        if st.button("➕ Nuevo Lead", use_container_width=True, type="primary"):
            st.session_state['crear_nuevo_lead'] = True
            st.rerun()
    
    # ========== APLICAR FILTROS ==========
    # Si los filtros están vacíos, mostrar todos
    if filtro_estado:
        df_filtrado = df_leads[df_leads['Estado Lead'].isin(filtro_estado)].copy()
    else:
        df_filtrado = df_leads.copy()
    
    if filtro_comercial:
        df_filtrado = df_filtrado[df_filtrado['Comercial Asignado'].isin(filtro_comercial)]
    
    if buscar:
        df_filtrado = df_filtrado[
            df_filtrado['Nombre Comercial'].str.contains(buscar, case=False, na=False)
        ]
    
    # ========== MÉTRICAS ==========
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        leads = len(df_leads[df_leads['Estado Lead'] == 'Lead'])
        st.metric("📋 Leads", leads)
    
    with col2:
        # Contar clientes activos directamente sin caché
        from openpyxl import load_workbook
        try:
            wb = load_workbook(config.ARCHIVO_CRM, data_only=True)
            ws = wb['CLIENTES_ACTIVOS']
            clientes = ws.max_row - 1  # Restar 1 por el encabezado
            wb.close()
        except:
            clientes = 0
        st.metric("✅ Clientes", clientes)
    
    with col3:
        bajas = len(df_leads[df_leads['Estado Lead'] == 'Baja'])
        st.metric("❌ Bajas", bajas)
    
    with col4:
        # Contar contactos directamente sin caché
        try:
            wb = load_workbook(config.ARCHIVO_CRM, data_only=True)
            ws = wb['CONTACTOS']
            total_contactos = ws.max_row - 1  # Restar 1 por el encabezado
            wb.close()
        except:
            total_contactos = 0
        st.metric("👥 Contactos", total_contactos)
    
    st.markdown("---")
    
    st.markdown("---")
    
    # ========== TABLA DE LEADS ==========
    if df_filtrado.empty:
        st.info("No se encontraron leads con los filtros aplicados.")
    else:
        # Botones de acción antes de la tabla
        col1, col2, col3 = st.columns([2, 2, 1])
        lead_a_convertir = None
        with col1:
            # Selector de lead para convertir
            leads_disponibles = df_filtrado[df_filtrado['Estado Lead'] == 'Lead']['Nombre Comercial'].tolist()
            if leads_disponibles:
                lead_a_convertir = st.selectbox(
                    "Selecciona un Lead para convertir a Cliente",
                    options=['-- Seleccionar --'] + leads_disponibles,
                    key='selector_lead_convertir'
                )
        with col2:
            if lead_a_convertir and lead_a_convertir != '-- Seleccionar --':
                if st.button("🎯 Convertir a Cliente", type="primary", use_container_width=True):
                    # Obtener datos del lead seleccionado
                    lead_seleccionado = df_filtrado[df_filtrado['Nombre Comercial'] == lead_a_convertir].iloc[0]
                    
                    st.session_state['convertir_a_cliente'] = True
                    st.session_state['id_lead_convertir'] = lead_seleccionado.get('ID')
                    st.session_state['nombre_lead_convertir'] = lead_seleccionado.get('Nombre Comercial', 'N/A')
                    st.session_state['datos_lead_convertir'] = lead_seleccionado.to_dict()
                    st.rerun()
        
        st.markdown("---")
        
        # Mostrar tabla con columnas seleccionadas (SIN edición de estado)
        columnas_mostrar = [
            'Nombre Comercial', 'Nombre Contacto', 'Email', 'Teléfono',
            'Estado Lead', 'Fuente Captación', 'Comercial Asignado'
        ]
        
        # Verificar que las columnas existan
        columnas_mostrar = [col for col in columnas_mostrar if col in df_filtrado.columns]
        
        df_tabla = df_filtrado[columnas_mostrar].copy()
        
        st.dataframe(
            df_tabla,
            use_container_width=True,
            hide_index=True,
            column_config={
                'Nombre Comercial': st.column_config.TextColumn('Cliente', width='medium'),
                'Nombre Contacto': st.column_config.TextColumn('Contacto', width='small'),
                'Email': st.column_config.TextColumn('Correo', width='medium'),
                'Teléfono': st.column_config.TextColumn('Teléfono', width='small'),
                'Estado Lead': st.column_config.TextColumn('Estado', width='small'),
                'Fuente Captación': st.column_config.TextColumn('Fuente', width='medium'),
                'Comercial Asignado': st.column_config.TextColumn('Responsable', width='medium')
            }
        )
        
        st.caption(f"Mostrando {len(df_filtrado)} de {len(df_leads)} leads")
    
    st.markdown("---")
    
    # Métricas
    col1, col2, col3, col4 = st.columns(4)
    
    st.markdown("---")
    
    # ========== MODAL NUEVO LEAD ==========
    if st.session_state.get('crear_nuevo_lead', False):
        crear_nuevo_lead_modal()
    
    # ========== MODAL CONVERTIR A CLIENTE ==========
    if st.session_state.get('convertir_a_cliente', False):
        convertir_lead_a_cliente_modal()
    
    # ========== MODAL EDITAR CLIENTE INCOMPLETO ==========
    if st.session_state.get('editar_cliente', False):
        editar_cliente_incompleto_modal()

# ============================================================================
# TAB 2: VISTA 360 DEL CLIENTE
# ============================================================================

def mostrar_vista_360_cliente():
    """Vista 360 profesional del cliente"""
    
    # Cargar clientes activos directamente sin caché
    from openpyxl import load_workbook
    try:
        wb = load_workbook(config.ARCHIVO_CRM, data_only=True)
        ws = wb['CLIENTES_ACTIVOS']
        
        datos = []
        encabezados = None
        
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                encabezados = row
            else:
                datos.append(row)
        
        if encabezados:
            df_clientes_activos = pd.DataFrame(datos, columns=encabezados)
        else:
            df_clientes_activos = pd.DataFrame()
        
        wb.close()
    except Exception as e:
        st.error(f"❌ Error al cargar CLIENTES_ACTIVOS: {e}")
        st.info("💡 Verifica que el archivo Excel existe y la hoja CLIENTES_ACTIVOS está creada.")
        return

    # Asegurar columnas mínimas para evitar KeyError
    columnas_minimas = ["Dirección", "Ciudad", "Nombre Comercial", "ID"]
    for col in columnas_minimas:
        if col not in df_clientes_activos.columns:
            df_clientes_activos[col] = ""
    
    # Filtrar clientes que tengan dirección Y ciudad (obligatorios para geolocalización)
    df_clientes_validos = df_clientes_activos[
        (df_clientes_activos['Dirección'].notna()) & 
        (df_clientes_activos['Dirección'] != '') &
        (df_clientes_activos['Ciudad'].notna()) & 
        (df_clientes_activos['Ciudad'] != '')
    ].copy()
    
    # Mostrar advertencia si hay clientes incompletos
    clientes_incompletos = df_clientes_activos[
        (df_clientes_activos['Dirección'].isna() | 
         (df_clientes_activos['Dirección'] == '') |
         df_clientes_activos['Ciudad'].isna() | 
         (df_clientes_activos['Ciudad'] == ''))
    ]
    
    if not clientes_incompletos.empty:
        with st.expander("⚠️ CLIENTES INCOMPLETOS (No aparecen en selector)"):
            st.warning(f"Hay {len(clientes_incompletos)} cliente(s) que no tienen Dirección o Ciudad completas:")
            
            for idx, row in clientes_incompletos.iterrows():
                col1, col2 = st.columns([4, 1])
                
                with col1:
                    nombre_cliente = row['Nombre Comercial']
                    falta = []
                    if pd.isna(row.get('Dirección')) or row.get('Dirección') == '':
                        falta.append("Dirección")
                    if pd.isna(row.get('Ciudad')) or row.get('Ciudad') == '':
                        falta.append("Ciudad")
                    
                    st.write(f"• **{nombre_cliente}** (Falta: {', '.join(falta)})")
                
                with col2:
                    if st.button("✏️ Editar", key=f"edit_incompleto_{idx}", use_container_width=True):
                        st.session_state['id_cliente_editar'] = row.get('ID')
                        st.session_state['nombre_cliente_editar'] = nombre_cliente
                        st.session_state['editar_cliente'] = True
                        st.rerun()
            
            st.info("💡 Haz click en ✏️ para rellenar Dirección + Ciudad")
    
    # Debug: mostrar información de la carga
    if df_clientes_validos.empty:
        st.warning("⚠️ No hay clientes activos completos (sin Dirección o Ciudad).")
        st.info("💡 Ve a la pestaña '📋 Gestión de Leads' y rellena Dirección + Ciudad para cada cliente.")
        
        # Mostrar información de debug
        with st.expander("🔧 Información de diagnóstico"):
            st.write(f"**Archivo:** {config.ARCHIVO_CRM}")
            st.write(f"**Hoja:** CLIENTES_ACTIVOS")
            st.write(f"**Filas totales:** {len(df_clientes_activos)}")
            st.write(f"**Filas válidas (con Dirección + Ciudad):** {len(df_clientes_validos)}")
            st.write(f"**Columnas:** {list(df_clientes_activos.columns) if not df_clientes_activos.empty else 'N/A'}")
        return
    
    # Verificar columna necesaria
    if 'Nombre Comercial' not in df_clientes_validos.columns:
        st.error("❌ Error: La hoja CLIENTES_ACTIVOS no tiene la columna 'Nombre Comercial'")
        st.info("💡 Ejecuta el script de migración o verifica la estructura del Excel.")
        with st.expander("🔧 Columnas encontradas"):
            st.write(list(df_clientes_validos.columns))
        return
    
    # Selector de cliente
    cliente_default = st.session_state.get('cliente_vista_360', None)
    
    # Verificar que el cliente seleccionado sea válido (en df_clientes_validos)
    if cliente_default and cliente_default in df_clientes_validos['Nombre Comercial'].values:
        idx_default = int(df_clientes_validos[df_clientes_validos['Nombre Comercial'] == cliente_default].index[0])
    else:
        idx_default = 0
    
    # Selector más visual
    st.markdown("### 🔍 Selecciona el Cliente")
    cliente_nombre = st.selectbox(
        "Cliente:",
        options=df_clientes_validos['Nombre Comercial'].tolist(),
        index=idx_default,
        key="selector_vista_360",
        label_visibility="collapsed"
    )
    
    # Obtener datos del cliente desde df_clientes_validos
    cliente = df_clientes_validos[df_clientes_validos['Nombre Comercial'] == cliente_nombre].iloc[0]
    id_cliente = cliente.get('ID', cliente_nombre)
    
    st.markdown("---")
    
    # ========== CABECERA DE IMPACTO: KPIs RÁPIDOS ==========
    mostrar_kpis_cabecera(cliente, id_cliente)
    
    st.markdown("---")
    
    # ========== BLOQUE CENTRAL: 360 DEL CLIENTE ==========
    mostrar_bloque_360(cliente, id_cliente, cliente_nombre)

def mostrar_kpis_cabecera(cliente, id_cliente):
    """Cabecera con KPIs - LTV calculado por servicios/cuotas activos"""
    
    col1, col2, col3, col4 = st.columns(4)
    
    # Limpiar caché y recargar datos frescos
    st.cache_data.clear()
    
    # Cargar datos de CLIENTES_ACTIVOS para información completa
    df_clientes_activos = utils.leer_excel_forzado(config.ARCHIVO_CRM, "CLIENTES_ACTIVOS")
    cliente_activo = None
    if not df_clientes_activos.empty:
        mask = df_clientes_activos['Nombre Comercial'] == cliente.get('Nombre Comercial')
        if mask.any():
            cliente_activo = df_clientes_activos[mask].iloc[0]
    
    # Calcular LTV (Lifetime Value) usando HISTORIAL_CONTRATOS + servicio actual
    #  - HISTORIAL_CONTRATOS: registros previos con 'Precio Anterior' y 'Tipo' ('Cuota' o 'Puntual')
    #  - Servicio actual: si es 'Cuota' se multiplica por meses activos; si es 'Puntual' se suma una vez
    ltv_historico = 0.0
    ltv_actual = 0.0

    try:
        df_hist = utils.leer_excel_forzado(config.ARCHIVO_CRM, "HISTORIAL_CONTRATOS")
    except:
        df_hist = pd.DataFrame()

    if not df_hist.empty and 'ID Cliente' in df_hist.columns:
        rows_hist = df_hist[df_hist['ID Cliente'] == id_cliente]
        for _, h in rows_hist.iterrows():
            try:
                precio_ant = h.get('Precio Anterior', 0) or 0
                tipo = h.get('Tipo', 'Cuota')
                fecha_inicio = h.get('Fecha Inicio')
                fecha_fin = h.get('Fecha Fin')

                # Normalizar fechas
                if pd.isna(fecha_inicio):
                    continue
                fi = pd.to_datetime(fecha_inicio)
                ff = pd.to_datetime(fecha_fin) if pd.notna(fecha_fin) and fecha_fin != '' else datetime.now()

                meses = max(1, (ff.year - fi.year) * 12 + (ff.month - fi.month))

                if tipo and str(tipo).lower().startswith('cuota'):
                    ltv_historico += float(precio_ant or 0) * meses
                else:
                    # Pago puntual o único
                    ltv_historico += float(precio_ant or 0)
            except Exception:
                continue

    # Calcular aportación del servicio actual
    if cliente_activo is not None:
        # Detectar tipo de facturación actual
        tipo_actual = cliente_activo.get('Tipo Facturación', '')
        precio_mensual = cliente_activo.get('Precio Mensual', 0) or 0
        precio_unico = cliente_activo.get('Precio Único', 0) or 0

        # Si no hay campo explícito, inferir por presencia de precio_unico
        if not tipo_actual:
            if precio_unico and precio_unico != 0:
                tipo_actual = 'Puntual'
            else:
                tipo_actual = 'Cuota'

        # Calcular meses activos desde Fecha Inicio
        meses_activo = 1
        if 'Fecha Inicio' in cliente_activo:
            try:
                fecha_ini = cliente_activo.get('Fecha Inicio')
                if pd.notna(fecha_ini) and fecha_ini != '':
                    fecha_ini_dt = pd.to_datetime(fecha_ini)
                    fecha_hoy = datetime.now()
                    meses_activo = (fecha_hoy.year - fecha_ini_dt.year) * 12 + (fecha_hoy.month - fecha_ini_dt.month)
                    if meses_activo == 0:
                        meses_activo = 1
                    elif fecha_hoy.day < fecha_ini_dt.day:
                        meses_activo = max(1, meses_activo)
            except:
                meses_activo = 1

        if str(tipo_actual).lower().startswith('cuota'):
            ltv_actual = float(precio_mensual or 0) * meses_activo
        else:
            ltv_actual = float(precio_unico or 0)

    ltv_total = ltv_historico + ltv_actual
    
    # ===== KPI 1: LTV ACUMULADO =====
    with col1:
        # Mostrar desglose histórico vs actual
        st.markdown(f"""
        <div style='background-color: #366092; padding: 18px; border-radius: 10px; text-align: left; min-height: 140px;'>
            <h4 style='margin: 0 0 8px 0; color: white; font-size: 14px;'>💰 LTV Acumulado</h4>
            <p style='margin: 0; color: white; font-weight: bold; font-size: 20px;'>{ltv_total:,.0f}€</p>
            <p style='margin: 6px 0 0 0; color: rgba(255,255,255,0.85); font-size: 12px;'>Histórico: {ltv_historico:,.0f}€ • Actual: {ltv_actual:,.0f}€</p>
        </div>
        """, unsafe_allow_html=True)
    
    # ===== KPI 2: SERVICIO ACTIVO =====
    with col2:
        # Servicio Activo - Desde CLIENTES_ACTIVOS
        if cliente_activo is not None and 'Servicio Contratado' in cliente_activo:
            servicio = cliente_activo.get('Servicio Contratado', 'No Asignado')
            if pd.isna(servicio):
                servicio = 'No Asignado'
        else:
            servicio = 'No Asignado'
        
        # Color según tipo de servicio
        if servicio == 'Completo':
            color_servicio = "#70AD47"  # Verde
        elif servicio == 'Cuota':
            color_servicio = "#5B9BD5"  # Azul
        elif servicio == 'Básico':
            color_servicio = "#FFC000"  # Amarillo
        else:
            color_servicio = "#999999"  # Gris
        
        st.markdown(f"""
        <div style='background-color: {color_servicio}; padding: 25px; border-radius: 10px; text-align: center; min-height: 150px; display: flex; flex-direction: column; justify-content: center;'>
            <h3 style='margin: 0; color: white; font-size: 28px;'>⚙️</h3>
            <h4 style='margin: 10px 0 5px 0; color: white; font-size: 14px;'>Servicio Activo</h4>
            <p style='margin: 5px 0 0 0; color: white; font-weight: bold; font-size: 18px;'>{servicio}</p>
        </div>
        """, unsafe_allow_html=True)
    
    # ===== KPI 3: ÚLTIMA INTERACCIÓN =====
    with col3:
        # Última Interacción
        df_int = utils.leer_excel_forzado(config.ARCHIVO_CRM, "INTERACCIONES")
        if not df_int.empty and 'ID Cliente' in df_int.columns:
            int_cliente = df_int[df_int['ID Cliente'] == id_cliente]
            if not int_cliente.empty:
                # Convertir Fecha a datetime si es necesario
                int_cliente['Fecha'] = pd.to_datetime(int_cliente['Fecha'], errors='coerce')
                ultima_fecha = int_cliente['Fecha'].max()
                dias = (datetime.now().date() - pd.to_datetime(ultima_fecha).date()).days
                texto_dias = f"Hace {dias} días"
                if dias > 30:
                    color_int = "#FF0000"
                elif dias > 14:
                    color_int = "#FFC000"
                else:
                    color_int = "#70AD47"
            else:
                texto_dias = "Sin interacciones"
                color_int = "#999999"
        else:
            texto_dias = "Sin interacciones"
            color_int = "#999999"
        
        st.markdown(f"""
        <div style='background-color: {color_int}; padding: 25px; border-radius: 10px; text-align: center; min-height: 150px; display: flex; flex-direction: column; justify-content: center;'>
            <h3 style='margin: 0; color: white; font-size: 28px;'>🕐</h3>
            <h4 style='margin: 10px 0 5px 0; color: white; font-size: 14px;'>Última Interacción</h4>
            <p style='margin: 5px 0 0 0; color: white; font-weight: bold; font-size: 18px;'>{texto_dias}</p>
        </div>
        """, unsafe_allow_html=True)
    
    # ===== KPI 4: ENCARGADO =====
    with col4:
        # Encargado Actual
        encargado = cliente.get('Encargado', 'No asignado')
        if pd.isna(encargado) or encargado == '':
            encargado = 'No asignado'
        
        st.markdown(f"""
        <div style='background-color: #5B9BD5; padding: 25px; border-radius: 10px; text-align: center; min-height: 150px; display: flex; flex-direction: column; justify-content: center;'>
            <h3 style='margin: 0; color: white; font-size: 28px;'>👤</h3>
            <h4 style='margin: 10px 0 5px 0; color: white; font-size: 14px;'>Encargado</h4>
            <p style='margin: 5px 0 0 0; color: white; font-weight: bold; font-size: 18px;'>{encargado}</p>
        </div>
        """, unsafe_allow_html=True)

def mostrar_bloque_360(cliente, id_cliente, cliente_nombre):
    """Bloque central con pestañas 360"""
    
    # Crear columnas: 70% para tabs, 30% para acciones
    col_main, col_sidebar = st.columns([7, 3])
    
    with col_main:
        # Pestañas del 360
        tab1, tab2, tab3, tab4 = st.tabs([
            "📊 Perfil Comercial",
            "🛠️ Servicios y Contratos",
            "👥 Directorio de Contactos",
            "📝 Timeline de Actividad"
        ])
        
        with tab1:
            mostrar_perfil_comercial(cliente, id_cliente)
        
        with tab2:
            mostrar_servicios_contratos(cliente, id_cliente, cliente_nombre)
        
        with tab3:
            mostrar_directorio_contactos(cliente, id_cliente, cliente_nombre)
        
        with tab4:
            mostrar_timeline_actividad(cliente, id_cliente, cliente_nombre)
    
    with col_sidebar:
        mostrar_acciones_rapidas(cliente, id_cliente, cliente_nombre)

def mostrar_perfil_comercial(cliente, id_cliente):
    """Pestaña: Perfil Comercial"""
    st.markdown("### 📊 Perfil Comercial")
    
    with st.form(key=f"form_perfil_{id_cliente}"):
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Datos Básicos**")
            nombre = st.text_input("Nombre Comercial:", value=cliente.get('Nombre Comercial', ''))
            cif = st.text_input("CIF:", value=cliente.get('CIF', ''))
            
            # Manejo seguro del índice de tipo de local
            tipo_local_actual = cliente.get('Tipo Local', '')
            if tipo_local_actual and tipo_local_actual in config.TIPOS_LOCAL:
                index_tipo_local = config.TIPOS_LOCAL.index(tipo_local_actual)
            else:
                index_tipo_local = 0
            
            tipo_local = st.selectbox("Tipo de Local:", config.TIPOS_LOCAL, index=index_tipo_local)
        
        with col2:
            st.markdown("**Gestión Comercial**")
            encargado = st.text_input("👤 Encargado Actual:", value=cliente.get('Encargado', ''), 
                                      placeholder="Nombre de quien es responsable en este momento")
            
            # Manejo seguro del índice de fuente
            fuente_cliente = cliente.get('Fuente', '')
            if fuente_cliente and fuente_cliente in config.FUENTES_CAPTACION:
                index_fuente = config.FUENTES_CAPTACION.index(fuente_cliente)
            else:
                index_fuente = 0
            
            fuente = st.selectbox("Fuente de Captación:", config.FUENTES_CAPTACION, index=index_fuente)
        
        # ===== SECCIÓN COMPLETA DE DIRECCIÓN =====
        st.markdown("### 📍 Dirección y Geolocalización")
        
        # Fila 1: Dirección Completa
        direccion = st.text_input(
            "Dirección Completa *",
            value=cliente.get('Dirección', ''),
            placeholder="Calle, número, piso...",
            help="Esta dirección se usará para mostrar el cliente en el mapa del Dashboard"
        )
        
        # Fila 2: Código Postal, Ciudad y Provincia
        col_cp, col_ciudad, col_provincia = st.columns(3)
        
        with col_cp:
            codigo_postal = st.text_input(
                "Código Postal",
                value=cliente.get('Código Postal', ''),
                placeholder="28001",
                max_chars=5
            )
        
        with col_ciudad:
            ciudad = st.text_input(
                "Ciudad *",
                value=cliente.get('Ciudad', ''),
                placeholder="Madrid, Barcelona...",
                help="Requerido para geolocalización correcta"
            )
        
        with col_provincia:
            provincia = st.text_input(
                "Provincia",
                value=cliente.get('Provincia', ''),
                placeholder="Madrid, Cataluña..."
            )
        
        col_t1, col_t2 = st.columns(2)
        with col_t1:
            telefono = st.text_input("Teléfono:", value=cliente.get('Teléfono', ''))
        with col_t2:
            email = st.text_input("Email:", value=cliente.get('Email', ''))
        
        notas = st.text_area("Notas / Temas de Interés:", value=cliente.get('Notas', ''), height=100)
        
        submitted = st.form_submit_button("💾 Guardar Cambios", type="primary", use_container_width=True)
        
        if submitted:
            # ===== VALIDACIÓN: CAMPOS OBLIGATORIOS =====
            errores = []
            
            # Dirección es obligatoria
            if not direccion or not direccion.strip():
                errores.append("📍 **Dirección Completa** es obligatoria para geolocalización")
            
            # Ciudad es obligatoria para el mapa
            if not ciudad or not ciudad.strip():
                errores.append("🏙️ **Ciudad** es obligatoria (necesaria para mostrar en el mapa)")
            
            # Si hay errores, mostrarlos y DETENER
            if errores:
                st.error("❌ **NO se puede guardar - Faltan campos obligatorios:**")
                for error in errores:
                    st.error(error)
                st.stop()  # Detener AQUÍ, ANTES de intentar guardar
            
            # Si llegamos aquí, la validación pasó
            with st.spinner("Guardando cambios..."):
                try:
                    # Limpiar caché y leer datos frescos
                    st.cache_data.clear()
                    import time
                    time.sleep(0.5)
                    
                    # ===== GUARDAR EN LEADS =====
                    df_clientes = utils.leer_excel_forzado(config.ARCHIVO_CRM, "LEADS")
                    if 'ID' in df_clientes.columns:
                        mascara = df_clientes['ID'] == id_cliente
                    else:
                        mascara = df_clientes['Nombre Comercial'] == cliente.get('Nombre Comercial')
                    
                    # Crear columnas de dirección si no existen en LEADS
                    for col in ['Dirección', 'Código Postal', 'Ciudad', 'Provincia', 'Teléfono', 'Email', 'Notas']:
                        if col not in df_clientes.columns:
                            df_clientes[col] = None
                    
                    # Actualizar TODOS los campos en LEADS
                    df_clientes.loc[mascara, 'Nombre Comercial'] = nombre
                    df_clientes.loc[mascara, 'CIF'] = cif
                    df_clientes.loc[mascara, 'Tipo Local'] = tipo_local
                    df_clientes.loc[mascara, 'Encargado'] = encargado
                    df_clientes.loc[mascara, 'Fuente'] = fuente
                    df_clientes.loc[mascara, 'Dirección'] = direccion
                    df_clientes.loc[mascara, 'Código Postal'] = codigo_postal
                    df_clientes.loc[mascara, 'Ciudad'] = ciudad
                    df_clientes.loc[mascara, 'Provincia'] = provincia
                    df_clientes.loc[mascara, 'Teléfono'] = str(telefono).strip() if telefono else ''
                    df_clientes.loc[mascara, 'Email'] = email
                    df_clientes.loc[mascara, 'Notas'] = notas
                    
                    # Escribir en LEADS con reintentos
                    guardado_leads = False
                    for intento in range(3):
                        try:
                            st.cache_data.clear()
                            if utils.escribir_excel(config.ARCHIVO_CRM, "LEADS", df_clientes):
                                guardado_leads = True
                                st.success("✅ Datos guardados en LEADS")
                                break
                        except Exception as e_leads:
                            if intento < 2:
                                time.sleep(1)
                            else:
                                raise e_leads
                    
                    if not guardado_leads:
                        st.error("❌ Error al guardar en LEADS después de reintentos")
                        st.stop()
                    
                    time.sleep(0.5)
                    
                    # ===== GUARDAR EN CLIENTES_ACTIVOS SI EXISTE =====
                    df_activos = utils.leer_excel_forzado(config.ARCHIVO_CRM, "CLIENTES_ACTIVOS")
                    if not df_activos.empty:
                        if 'ID' in df_activos.columns:
                            mascara_activos = df_activos['ID'] == id_cliente
                        else:
                            mascara_activos = df_activos['Nombre Comercial'] == nombre

                        if mascara_activos.any():
                            # Crear todas las columnas necesarias si no existen
                            for col in ['Encargado', 'Código Postal', 'Provincia', 'Latitud', 'Longitud', 'Dirección', 'Ciudad', 'Teléfono', 'Email', 'Nombre Comercial', 'CIF', 'Tipo Local']:
                                if col not in df_activos.columns:
                                    df_activos[col] = None
                            
                            # Geocodificar la dirección REALMENTE con Nominatim
                            lat, lon = None, None
                            
                            # Intentar con Nominatim para geocodificación real
                            try:
                                from geopy.geocoders import Nominatim
                                geolocator = Nominatim(user_agent="horeca_crm", timeout=5)
                                direccion_completa = f"{direccion.strip()}, {ciudad.strip()}, España"
                                
                                try:
                                    location = geolocator.geocode(direccion_completa, language='es')
                                    if location:
                                        lat = location.latitude
                                        lon = location.longitude
                                    else:
                                        # Si Nominatim no encuentra, usar caché
                                        ciudades_cache = {
                                            'madrid': (40.4168, -3.7038),
                                            'barcelona': (41.3874, 2.1686),
                                            'valencia': (39.4699, -0.3763),
                                            'sevilla': (37.3891, -5.9845),
                                            'bilbao': (43.2630, -2.9350),
                                            'malaga': (36.7213, -4.4214),
                                            'murcia': (37.9922, -1.1307),
                                            'zaragoza': (41.6488, -0.8891),
                                            'palma': (39.5696, 2.6502),
                                            'pamplona': (42.8125, -1.6458),
                                            'alicante': (38.3452, -0.4810),
                                            'toledo': (39.8628, -4.0273),
                                            'segovia': (40.9526, -4.1197),
                                            'alcalá de henares': (40.4819, -3.3589),
                                            'getafe': (40.3078, -3.7251),
                                            'leganés': (40.3295, -3.7566),
                                            'fuenlabrada': (40.2900, -3.8100),
                                            'móstoles': (40.3239, -3.8623),
                                            'rivas-vaciamadrid': (40.1761, -3.5223),
                                        }
                                        ciudad_lower = ciudad.lower().strip() if isinstance(ciudad, str) and ciudad.strip() else ''
                                        lat, lon = ciudades_cache.get(ciudad_lower, (40.4168, -3.7038))
                                except Exception as e_nominatim:
                                    # Fallback a caché si hay error
                                    ciudades_cache = {
                                        'madrid': (40.4168, -3.7038),
                                        'barcelona': (41.3874, 2.1686),
                                        'valencia': (39.4699, -0.3763),
                                        'sevilla': (37.3891, -5.9845),
                                        'bilbao': (43.2630, -2.9350),
                                        'malaga': (36.7213, -4.4214),
                                        'murcia': (37.9922, -1.1307),
                                        'zaragoza': (41.6488, -0.8891),
                                        'palma': (39.5696, 2.6502),
                                        'pamplona': (42.8125, -1.6458),
                                        'alicante': (38.3452, -0.4810),
                                        'toledo': (39.8628, -4.0273),
                                        'segovia': (40.9526, -4.1197),
                                        'alcalá de henares': (40.4819, -3.3589),
                                        'getafe': (40.3078, -3.7251),
                                        'leganés': (40.3295, -3.7566),
                                        'fuenlabrada': (40.2900, -3.8100),
                                        'móstoles': (40.3239, -3.8623),
                                        'rivas-vaciamadrid': (40.1761, -3.5223),
                                    }
                                    ciudad_lower = ciudad.lower().strip() if isinstance(ciudad, str) and ciudad.strip() else ''
                                    lat, lon = ciudades_cache.get(ciudad_lower, (40.4168, -3.7038))
                            except ImportError:
                                # Si no tiene geopy, usar caché
                                ciudades_cache = {
                                    'madrid': (40.4168, -3.7038),
                                    'barcelona': (41.3874, 2.1686),
                                    'valencia': (39.4699, -0.3763),
                                    'sevilla': (37.3891, -5.9845),
                                    'bilbao': (43.2630, -2.9350),
                                    'malaga': (36.7213, -4.4214),
                                    'murcia': (37.9922, -1.1307),
                                    'zaragoza': (41.6488, -0.8891),
                                    'palma': (39.5696, 2.6502),
                                    'pamplona': (42.8125, -1.6458),
                                    'alicante': (38.3452, -0.4810),
                                    'toledo': (39.8628, -4.0273),
                                    'segovia': (40.9526, -4.1197),
                                    'alcalá de henares': (40.4819, -3.3589),
                                    'getafe': (40.3078, -3.7251),
                                    'leganés': (40.3295, -3.7566),
                                    'fuenlabrada': (40.2900, -3.8100),
                                    'móstoles': (40.3239, -3.8623),
                                    'rivas-vaciamadrid': (40.1761, -3.5223),
                                }
                                ciudad_lower = ciudad.lower().strip() if isinstance(ciudad, str) and ciudad.strip() else ''
                                lat, lon = ciudades_cache.get(ciudad_lower, (40.4168, -3.7038))
                            
                            # Actualizar TODOS los campos (CON FLOAT CONVERSION PARA LAT/LON)
                            df_activos.loc[mascara_activos, 'Encargado'] = encargado
                            df_activos.loc[mascara_activos, 'Nombre Comercial'] = nombre
                            df_activos.loc[mascara_activos, 'CIF'] = cif
                            df_activos.loc[mascara_activos, 'Tipo Local'] = tipo_local
                            df_activos.loc[mascara_activos, 'Dirección'] = direccion
                            df_activos.loc[mascara_activos, 'Código Postal'] = codigo_postal
                            df_activos.loc[mascara_activos, 'Ciudad'] = ciudad
                            df_activos.loc[mascara_activos, 'Provincia'] = provincia
                            df_activos.loc[mascara_activos, 'Latitud'] = float(lat) if lat is not None else None
                            df_activos.loc[mascara_activos, 'Longitud'] = float(lon) if lon is not None else None
                            df_activos.loc[mascara_activos, 'Teléfono'] = str(telefono).strip() if telefono else ''
                            df_activos.loc[mascara_activos, 'Email'] = email
                            
                            # Escribir en CLIENTES_ACTIVOS con reintentos
                            guardado_activos = False
                            for intento in range(3):
                                try:
                                    st.cache_data.clear()
                                    if utils.escribir_excel(config.ARCHIVO_CRM, "CLIENTES_ACTIVOS", df_activos):
                                        guardado_activos = True
                                        st.success("✅ Datos guardados en CLIENTES_ACTIVOS")
                                        break
                                except Exception as e_activos:
                                    if intento < 2:
                                        time.sleep(1)
                                    else:
                                        raise e_activos
                            
                            if not guardado_activos:
                                st.error("❌ Error al guardar en CLIENTES_ACTIVOS después de reintentos")
                    
                    # Éxito final
                    st.success("✅ ¡Perfil actualizado correctamente!")
                    st.cache_data.clear()
                    time.sleep(1)
                    st.rerun()
                
                except Exception as e:
                    import traceback
                    st.error(f"❌ Error durante el guardado: {str(e)}")
                    st.error(f"Detalles: {traceback.format_exc()}")

def mostrar_servicios_contratos(cliente, id_cliente, cliente_nombre):
    """Pestaña: Servicios y Contratos - MEJORADA CON DATOS REALES"""
    st.markdown("### 🛠️ Servicios y Contratos")
    
    # Cargar datos del cliente activo
    df_clientes_activos = utils.leer_excel_forzado(config.ARCHIVO_CRM, "CLIENTES_ACTIVOS")
    
    if not df_clientes_activos.empty:
        mask = df_clientes_activos['Nombre Comercial'] == cliente_nombre
        if mask.any():
            cliente_activo = df_clientes_activos[mask].iloc[0]
            
            # Mostrar servicio contratado
            servicio_contratado = cliente_activo.get('Servicio Contratado', 'No definido')
            precio_mensual = cliente_activo.get('Precio Mensual', 0)
            if pd.isna(precio_mensual):
                precio_mensual = 0
            
            fecha_inicio = cliente_activo.get('Fecha Inicio', 'N/A')
            estado_servicio = cliente_activo.get('Estado', 'N/A')
            
            # Tarjeta del servicio actual
            with st.container():
                st.markdown(f"""
                <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                            padding: 25px; border-radius: 15px; margin-bottom: 20px;'>
                    <h3 style='color: white; margin: 0;'>📦 {servicio_contratado}</h3>
                    <p style='color: white; margin: 5px 0; opacity: 0.9;'>Estado: {estado_servicio}</p>
                </div>
                """, unsafe_allow_html=True)
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.metric("💰 Precio Mensual", f"{precio_mensual:.0f}€")
                
                with col2:
                    # Mostrar solo la fecha, no la hora
                    if isinstance(fecha_inicio, str):
                        fecha_str = fecha_inicio.split()[0] if ' ' in fecha_inicio else fecha_inicio
                    else:
                        fecha_str = pd.to_datetime(fecha_inicio).strftime("%d/%m/%Y") if pd.notna(fecha_inicio) else "N/A"
                    st.write(f"**📅 Inicio:** {fecha_str}")
            
            st.markdown("---")
            
            # Opción para cambiar servicio
            st.markdown("### Servicio Contratado")
            
            with st.expander("✏️ Modificar Servicio Contratado"):
                with st.form(key=f"form_cambiar_servicio_{id_cliente}"):
                    nuevo_servicio = st.selectbox(
                        "Nuevo Servicio:",
                        options=list(config.SERVICIOS_DISPONIBLES.keys()),
                        index=list(config.SERVICIOS_DISPONIBLES.keys()).index(servicio_contratado) if servicio_contratado in config.SERVICIOS_DISPONIBLES else 0
                    )

                    # Tipo de facturación: Cuota (mensual) o Pago único
                    tipo_fact = st.selectbox("Tipo de Facturación:", options=['Cuota', 'Puntual'], index=0)

                    precio_info = config.SERVICIOS_DISPONIBLES[nuevo_servicio]
                    st.info(f"💰 **Precio sugerido:** {precio_info['precio']}€ - {precio_info['descripcion']}")

                    # Mostrar campos según tipo
                    if tipo_fact == 'Cuota':
                        nuevo_precio_mensual = st.number_input("Precio Mensual (€):", value=float(precio_info['precio']), min_value=0.0, step=1.0)
                        nuevo_precio_unico = 0
                    else:
                        nuevo_precio_unico = st.number_input("Precio Único (€):", value=0.0, min_value=0.0, step=1.0)
                        nuevo_precio_mensual = 0

                    nueva_fecha_inicio = st.date_input("Fecha Inicio:", value=pd.to_datetime(fecha_inicio) if pd.notna(fecha_inicio) else datetime.now())
                    nuevo_estado = st.selectbox("Estado:", ['Activo', 'Pausado', 'Cancelado'])

                    if st.form_submit_button("💾 Guardar Cambios", type="primary"):
                        with st.spinner("Actualizando servicio y guardando historial..."):
                            # Limpiar caché y leer datos frescos
                            st.cache_data.clear()
                            time.sleep(0.5)
                            
                            # Preparar historial: guardar el servicio anterior si existía
                            try:
                                df_hist = utils.leer_excel_forzado(config.ARCHIVO_CRM, 'HISTORIAL_CONTRATOS')
                            except:
                                df_hist = pd.DataFrame()

                            prev_serv = cliente_activo.get('Servicio Contratado', '') if cliente_activo is not None else ''
                            prev_prec_m = cliente_activo.get('Precio Mensual', 0) if cliente_activo is not None else 0
                            prev_prec_u = cliente_activo.get('Precio Único', 0) if cliente_activo is not None else 0
                            prev_tipo = cliente_activo.get('Tipo Facturación', 'Cuota') if cliente_activo is not None else 'Cuota'
                            prev_fecha_inicio = cliente_activo.get('Fecha Inicio', None) if cliente_activo is not None else None

                            # Solo guardar historial si hay cambio de servicio
                            cambio_servicio = prev_serv != nuevo_servicio
                            
                            if prev_serv and (prev_prec_m or prev_prec_u) and cambio_servicio:
                                # Crear registro histórico
                                if df_hist.empty:
                                    df_hist = pd.DataFrame(columns=['ID','ID Cliente','Servicio Anterior','Precio Anterior','Fecha Inicio','Fecha Fin','Tipo','Motivo Cambio'])

                                precio_anterior = prev_prec_m if prev_tipo == 'Cuota' else prev_prec_u
                                nuevo_hist = {
                                    'ID': utils.generar_id(),
                                    'ID Cliente': id_cliente,
                                    'Servicio Anterior': prev_serv,
                                    'Precio Anterior': float(precio_anterior),
                                    'Fecha Inicio': prev_fecha_inicio,
                                    'Fecha Fin': nueva_fecha_inicio,
                                    'Tipo': prev_tipo,
                                    'Motivo Cambio': f'Cambio de {prev_serv} a {nuevo_servicio}'
                                }

                                df_hist = pd.concat([df_hist, pd.DataFrame([nuevo_hist])], ignore_index=True)
                                
                                if not utils.escribir_excel(config.ARCHIVO_CRM, 'HISTORIAL_CONTRATOS', df_hist):
                                    st.error("❌ Error al guardar historial de contratos")
                                    return
                                
                                time.sleep(1)
                                st.success(f"📝 Servicio anterior '{prev_serv}' guardado en historial")

                            # Actualizar en CLIENTES_ACTIVOS
                            df_clientes_activos_fresh = utils.leer_excel_forzado(config.ARCHIVO_CRM, "CLIENTES_ACTIVOS")
                            mask_fresh = df_clientes_activos_fresh['Nombre Comercial'] == cliente_nombre
                            
                            df_clientes_activos_fresh.loc[mask_fresh, 'Servicio Contratado'] = nuevo_servicio
                            df_clientes_activos_fresh.loc[mask_fresh, 'Precio Mensual'] = nuevo_precio_mensual
                            df_clientes_activos_fresh.loc[mask_fresh, 'Precio Único'] = nuevo_precio_unico
                            df_clientes_activos_fresh.loc[mask_fresh, 'Tipo Facturación'] = tipo_fact
                            df_clientes_activos_fresh.loc[mask_fresh, 'Fecha Inicio'] = nueva_fecha_inicio
                            df_clientes_activos_fresh.loc[mask_fresh, 'Estado'] = nuevo_estado
                            df_clientes_activos_fresh.loc[mask_fresh, 'MRR'] = nuevo_precio_mensual if nuevo_estado == 'Activo' and tipo_fact == 'Cuota' else 0

                            # Escribir cambios
                            st.cache_data.clear()
                            if not utils.escribir_excel(config.ARCHIVO_CRM, "CLIENTES_ACTIVOS", df_clientes_activos_fresh):
                                st.error("❌ Error al actualizar servicio en CLIENTES_ACTIVOS")
                                return
                            
                            time.sleep(1)
                            st.success("✅ Servicio actualizado correctamente")
                            st.cache_data.clear()
                            time.sleep(0.5)
                            st.rerun()
            
            # Mostrar historial de contratos
            st.markdown("---")
            st.markdown("### 📜 Historial de Servicios Contratados")
            
            try:
                df_hist = utils.leer_excel_forzado(config.ARCHIVO_CRM, 'HISTORIAL_CONTRATOS')
                if not df_hist.empty and 'ID Cliente' in df_hist.columns:
                    # Convertir Fecha Fin a datetime
                    df_hist['Fecha Fin'] = pd.to_datetime(df_hist['Fecha Fin'], errors='coerce')
                    hist_cliente = df_hist[df_hist['ID Cliente'] == id_cliente].sort_values('Fecha Fin', ascending=False)
                    
                    if not hist_cliente.empty:
                        st.info(f"📊 Total de cambios registrados: **{len(hist_cliente)}**")
                        
                        for idx, (hist_idx, hist) in enumerate(hist_cliente.iterrows()):
                            servicio = hist.get('Servicio Anterior', 'N/A')
                            precio = hist.get('Precio Anterior', 0)
                            tipo = hist.get('Tipo', 'Cuota')
                            fecha_inicio = hist.get('Fecha Inicio')
                            fecha_fin = hist.get('Fecha Fin')
                            motivo = hist.get('Motivo Cambio', 'N/A')
                            
                            # Calcular meses
                            try:
                                if pd.notna(fecha_inicio) and pd.notna(fecha_fin):
                                    fi = pd.to_datetime(fecha_inicio)
                                    ff = pd.to_datetime(fecha_fin)
                                    meses = max(1, (ff.year - fi.year) * 12 + (ff.month - fi.month))
                                    
                                    # Calcular ingresos del periodo
                                    if tipo and str(tipo).lower().startswith('cuota'):
                                        ingresos_periodo = float(precio) * meses
                                        texto_periodo = f"{meses} meses × {precio:.0f}€/mes = {ingresos_periodo:.0f}€"
                                    else:
                                        ingresos_periodo = float(precio)
                                        texto_periodo = f"Pago único: {precio:.0f}€"
                                else:
                                    meses = 0
                                    texto_periodo = "Periodo no definido"
                                    ingresos_periodo = 0
                            except:
                                meses = 0
                                texto_periodo = "Error en cálculo"
                                ingresos_periodo = 0
                            
                            # Formatear fechas
                            fecha_i_str = pd.to_datetime(fecha_inicio).strftime("%d/%m/%Y") if pd.notna(fecha_inicio) else "N/A"
                            fecha_f_str = pd.to_datetime(fecha_fin).strftime("%d/%m/%Y") if pd.notna(fecha_fin) else "N/A"
                            
                            with st.expander(f"📦 {servicio} ({fecha_i_str} → {fecha_f_str})"):
                                col1, col2, col3, col4 = st.columns([1.5, 1.5, 0.5, 0.5])
                                
                                with col1:
                                    st.write(f"**Servicio:** {servicio}")
                                    st.write(f"**Tipo:** {tipo}")
                                    st.write(f"**Precio:** {precio:.0f}€")
                                
                                with col2:
                                    st.write(f"**Inicio:** {fecha_i_str}")
                                    st.write(f"**Fin:** {fecha_f_str}")
                                    st.write(f"**Meses activo:** {meses}")
                                
                                with col3:
                                    if st.button("✏️", key=f"edit_hist_{hist_idx}", help="Editar"):
                                        st.session_state[f"edit_mode_{hist_idx}"] = True
                                
                                with col4:
                                    if st.button("🗑️", key=f"del_hist_{hist_idx}", help="Eliminar"):
                                        # Eliminar directamente
                                        df_hist_updated = df_hist.drop(hist_idx)
                                        if utils.escribir_excel(config.ARCHIVO_CRM, "HISTORIAL_CONTRATOS", df_hist_updated):
                                            st.success("✅ Registro eliminado")
                                            st.cache_data.clear()
                                            time.sleep(0.5)
                                            st.rerun()
                                        else:
                                            st.error("❌ Error al eliminar")
                                
                                # Modo edición (mostrar si está activado)
                                if st.session_state.get(f"edit_mode_{hist_idx}"):
                                    st.divider()
                                    st.markdown("**✏️ Editar registro**")
                                    
                                    col_e1, col_e2, col_e3 = st.columns(3)
                                    with col_e1:
                                        servicio_edit = st.text_input("Servicio:", value=servicio, key=f"serv_edit_{hist_idx}")
                                        precio_edit = st.number_input("Precio:", value=float(precio), key=f"precio_edit_{hist_idx}")
                                    
                                    with col_e2:
                                        tipo_edit = st.selectbox("Tipo:", ["Cuota", "Puntual"], index=0 if str(tipo).startswith("Cuota") else 1, key=f"tipo_edit_{hist_idx}")
                                        motivo_edit = st.text_input("Motivo cambio:", value=str(motivo), key=f"motivo_edit_{hist_idx}")
                                    
                                    with col_e3:
                                        fecha_i_edit = st.date_input("Fecha inicio:", value=pd.to_datetime(fecha_inicio).date() if pd.notna(fecha_inicio) else datetime.now().date(), key=f"fi_edit_{hist_idx}")
                                        fecha_f_edit = st.date_input("Fecha fin:", value=pd.to_datetime(fecha_fin).date() if pd.notna(fecha_fin) else datetime.now().date(), key=f"ff_edit_{hist_idx}")
                                    
                                    col_save, col_cancel = st.columns(2)
                                    with col_save:
                                        if st.button("💾 Guardar", key=f"save_hist_{hist_idx}"):
                                            # Actualizar registro
                                            df_hist.loc[hist_idx, 'Servicio Anterior'] = servicio_edit
                                            df_hist.loc[hist_idx, 'Precio Anterior'] = precio_edit
                                            df_hist.loc[hist_idx, 'Tipo'] = tipo_edit
                                            df_hist.loc[hist_idx, 'Motivo Cambio'] = motivo_edit
                                            df_hist.loc[hist_idx, 'Fecha Inicio'] = pd.to_datetime(fecha_i_edit)
                                            df_hist.loc[hist_idx, 'Fecha Fin'] = pd.to_datetime(fecha_f_edit)
                                            
                                            if utils.escribir_excel(config.ARCHIVO_CRM, "HISTORIAL_CONTRATOS", df_hist):
                                                st.success("✅ Registro actualizado")
                                                st.cache_data.clear()
                                                time.sleep(0.5)
                                                st.rerun()
                                            else:
                                                st.error("❌ Error al guardar")
                                    
                                    with col_cancel:
                                        if st.button("❌ Cancelar", key=f"cancel_hist_{hist_idx}"):
                                            st.session_state[f"edit_mode_{hist_idx}"] = False
                                            st.rerun()
                                
                                st.info(f"💰 {texto_periodo}")
                                st.caption(f"📝 {motivo}")
                    else:
                        st.info("No hay historial de cambios de servicio para este cliente.")
                else:
                    st.info("No hay historial de cambios registrado.")
            except Exception as e:
                st.warning(f"⚠️ No se pudo cargar el historial: {e}")
                import traceback
                st.write(traceback.format_exc())
        else:
            st.warning("⚠️ Este cliente no está en la hoja CLIENTES_ACTIVOS")
            st.info("💡 Para gestionar servicios, el cliente debe estar registrado como activo.")
    else:
        st.warning("⚠️ No se pudo cargar la información de clientes activos")

def mostrar_directorio_contactos(cliente, id_cliente, cliente_nombre):
    """Pestaña: Directorio de Contactos - DESARROLLADA COMPLETAMENTE"""
    st.markdown("### 👥 Directorio de Contactos")
    
    # Cargar contactos del cliente
    try:
        df_contactos = utils.leer_excel_forzado(config.ARCHIVO_CRM, "CONTACTOS")
    except:
        df_contactos = pd.DataFrame()
    
    # Crear hoja CONTACTOS si no existe
    if df_contactos.empty:
        df_contactos = pd.DataFrame(columns=[
            'ID', 'ID Cliente', 'Nombre Cliente', 'Nombre Contacto', 
            'Cargo', 'Teléfono', 'Email', 'Empresa',
            'Es Principal', 'Notas', 'Fecha Registro'
        ])
    
    # Filtrar contactos de este cliente
    contactos_cliente = df_contactos[df_contactos['ID Cliente'] == id_cliente] if not df_contactos.empty and 'ID Cliente' in df_contactos.columns else pd.DataFrame()
    
    # Botón para añadir contacto
    col1, col2 = st.columns([3, 1])
    with col2:
        if st.button("➕ Añadir Contacto", use_container_width=True, type="primary"):
            st.session_state['crear_contacto'] = True
            st.session_state['id_cliente_contacto'] = id_cliente
            st.session_state['nombre_cliente_contacto'] = cliente_nombre
    
    st.markdown("---")
    
    # Mostrar contactos existentes
    if not contactos_cliente.empty:
        for idx, (_, contacto) in enumerate(contactos_cliente.iterrows()):
            es_principal = contacto.get('Es Principal', False)
            
            with st.expander(
                f"{'⭐ ' if es_principal else ''}👤 {contacto.get('Nombre Contacto', 'N/A')} - {contacto.get('Cargo', 'N/A')}", 
                expanded=es_principal
            ):
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.write(f"**Cargo:** {contacto.get('Cargo', 'N/A')}")
                    st.write(f"**Teléfono:** {contacto.get('Teléfono', 'N/A')}")
                
                with col2:
                    st.write(f"**Email:** {contacto.get('Email', 'N/A')}")
                    st.write(f"**Empresa:** {contacto.get('Empresa', 'N/A')}")
                
                with col3:
                    st.write(f"**Principal:** {'Sí' if es_principal else 'No'}")
                    fecha_reg = contacto.get('Fecha Registro', 'N/A')
                    st.write(f"**Registro:** {fecha_reg}")
                
                if pd.notna(contacto.get('Notas', '')) and contacto.get('Notas', '') != '':
                    st.info(f"📝 {contacto.get('Notas', '')}")
                
                # Botones de acción
                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    if st.button("✏️ Editar", key=f"edit_contacto_{idx}", use_container_width=True):
                        st.session_state['editar_contacto'] = True
                        st.session_state['contacto_a_editar'] = contacto
                        st.rerun()
                
                with col_btn2:
                    if st.button("🗑️ Eliminar", key=f"del_contacto_{idx}", use_container_width=True):
                        # Eliminar contacto
                        df_contactos = df_contactos[df_contactos['ID'] != contacto.get('ID')]
                        if utils.escribir_excel(config.ARCHIVO_CRM, "CONTACTOS", df_contactos):
                            st.success("✅ Contacto eliminado")
                            st.cache_data.clear()
                            time.sleep(0.5)
                            st.rerun()
    else:
        st.info("📝 No hay contactos registrados para este cliente. Añade el primero.")
    
    # Modal para crear contacto
    if st.session_state.get('crear_contacto') and st.session_state.get('id_cliente_contacto') == id_cliente:
        crear_contacto_modal(id_cliente, cliente_nombre)
    
    # Modal para editar contacto
    if st.session_state.get('editar_contacto'):
        editar_contacto_modal(st.session_state.get('contacto_a_editar', {}), id_cliente)

def mostrar_timeline_actividad(cliente, id_cliente, cliente_nombre):
    """Pestaña: Timeline de Actividad - CON ELIMINACIÓN"""
    st.markdown("### 📝 Timeline de Actividad")
    
    # Cargar interacciones
    df_int = utils.leer_excel_forzado(config.ARCHIVO_CRM, "INTERACCIONES")
    
    if not df_int.empty and 'ID Cliente' in df_int.columns:
        # Convertir Fecha a datetime
        df_int['Fecha'] = pd.to_datetime(df_int['Fecha'], errors='coerce')
        int_cliente = df_int[df_int['ID Cliente'] == id_cliente].sort_values('Fecha', ascending=False)
        
        if not int_cliente.empty:
            for idx, (row_idx, interaccion) in enumerate(int_cliente.iterrows()):
                fecha = pd.to_datetime(interaccion.get('Fecha')).strftime('%d/%m/%Y %H:%M')
                tipo = interaccion.get('Tipo', 'N/A')
                desc = interaccion.get('Descripción', '')
                
                # Iconos por tipo
                iconos = {
                    'Llamada': '📞',
                    'Email': '📧',
                    'Reunión': '🤝',
                    'WhatsApp': '💬',
                    'Visita': '🚗',
                    'Otro': '📝'
                }
                icono = iconos.get(tipo, '📝')
                
                # Contenedor con botón de eliminar
                col_main, col_btn = st.columns([9, 1])
                
                with col_main:
                    st.markdown(f"""
                    <div style='background-color: #f0f2f6; padding: 10px; border-radius: 5px; 
                                margin-bottom: 10px; border-left: 4px solid #366092;'>
                        <p style='margin: 0;'><strong>{icono} {tipo}</strong> - <em>{fecha}</em></p>
                        <p style='margin: 5px 0 0 0;'>{desc}</p>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col_btn:
                    if st.button("🗑️", key=f"del_int_{idx}", help="Eliminar interacción"):
                        # Eliminar interacción
                        df_int = df_int.drop(row_idx)
                        if utils.escribir_excel(config.ARCHIVO_CRM, "INTERACCIONES", df_int):
                            st.success("✅ Interacción eliminada")
                            st.cache_data.clear()
                            time.sleep(0.5)
                            st.rerun()
                        else:
                            st.error("❌ Error al eliminar")
        else:
            st.info("No hay interacciones registradas")
    else:
        st.info("No hay interacciones registradas")

def mostrar_acciones_rapidas(cliente, id_cliente, cliente_nombre):
    """Panel lateral con acciones rápidas"""
    st.markdown("### ⚡ ACCIONES RÁPIDAS")
    
    # Registrar Llamada/Visita
    if st.button("📞 Registrar Llamada/Visita", use_container_width=True, type="primary"):
        st.session_state['crear_interaccion'] = True
        st.session_state['id_cliente_interaccion'] = id_cliente
        st.session_state['nombre_cliente_interaccion'] = cliente_nombre
        st.rerun()
    
    # Enviar WhatsApp
    if st.button("📱 Enviar WhatsApp", use_container_width=True, type="secondary"):
        st.session_state['enviar_whatsapp_modal'] = True
        st.session_state['id_cliente_whatsapp'] = id_cliente
        st.session_state['nombre_cliente_whatsapp'] = cliente_nombre
        st.rerun()
    
    # Programar Tarea - AHORA FUNCIONAL
    if st.button("📅 Programar Tarea", use_container_width=True):
        st.session_state['programar_tarea'] = True
        st.session_state['id_cliente_tarea'] = id_cliente
        st.session_state['nombre_cliente_tarea'] = cliente_nombre
        st.rerun()
    
    # Subir Documento
    if st.button("📄 Subir Documento", use_container_width=True):
        st.info("🔧 Funcionalidad próximamente")
    
    # (Botón de incidencia eliminado por limpieza de interfaz)
    
    st.markdown("---")
    # Modal de crear interacción
    if st.session_state.get('crear_interaccion') and st.session_state.get('id_cliente_interaccion') == id_cliente:
        crear_interaccion_modal(id_cliente, cliente_nombre)
    
    # Modal de enviar WhatsApp
    if st.session_state.get('enviar_whatsapp_modal') and st.session_state.get('id_cliente_whatsapp') == id_cliente:
        crear_modal_whatsapp(id_cliente, cliente_nombre)
    
    # Modal de programar tarea
    if st.session_state.get('programar_tarea') and st.session_state.get('id_cliente_tarea') == id_cliente:
        programar_tarea_modal(id_cliente, cliente_nombre)

# ============================================================================
# TAB 3: PRÓXIMAS ACCIONES
# ============================================================================

def mostrar_proximas_acciones():
    """Vista de próximas acciones"""
    st.subheader("📅 Próximas Acciones")
    
    # Botón de refrescar
    col1, col2, col3 = st.columns([5, 1, 1])
    with col3:
        if st.button("🔄 Refrescar", use_container_width=True):
            st.cache_data.clear()
            st.rerun()
    
    # Cargar interacciones
    try:
        df_int = utils.leer_excel_forzado(config.ARCHIVO_CRM, "INTERACCIONES")
    except:
        df_int = pd.DataFrame()
    
    if df_int.empty:
        st.info("No hay acciones programadas.")
        return
    
    # Asegurar que la columna existe
    if 'Próxima Acción' not in df_int.columns:
        st.warning("⚠️ La hoja INTERACCIONES no tiene la columna 'Próxima Acción'")
        return
    
    # Filtrar con próxima acción
    df_acciones = df_int[
        (df_int['Próxima Acción'].notna()) & 
        (df_int['Próxima Acción'] != '')
    ].copy()
    
    if df_acciones.empty:
        st.info("No hay próximas acciones programadas.")
        return
    
    st.info(f"📊 Total: **{len(df_acciones)}** acciones")
    
    # Calcular urgencia
    hoy = datetime.now().date()
    
    def calcular_urgencia(fecha):
        if pd.isna(fecha):
            return 4
        try:
            fecha_dt = pd.to_datetime(fecha).date()
            dias = (fecha_dt - hoy).days
            if dias < 0:
                return 1
            elif dias == 0:
                return 2
            elif dias <= 7:
                return 3
            else:
                return 4
        except:
            return 4
    
    df_acciones['Urgencia'] = df_acciones['Fecha Próxima Acción'].apply(calcular_urgencia)
    df_acciones = df_acciones.sort_values('Urgencia')
    
    # Mostrar por categorías
    st.markdown("---")
    st.markdown("### 🔴 URGENTE (Vencidas)")
    vencidas = df_acciones[df_acciones['Urgencia'] == 1]
    if not vencidas.empty:
        for idx, (row_idx, row) in enumerate(vencidas.iterrows()):
            try:
                fecha = pd.to_datetime(row.get('Fecha Próxima Acción')).strftime('%d/%m/%Y')
            except:
                fecha = str(row.get('Fecha Próxima Acción', 'N/A'))
            
            col1, col2 = st.columns([9, 1])
            with col1:
                st.error(f"⚠️ **{fecha}** - {row.get('Nombre Cliente', 'N/A')} - {row.get('Próxima Acción', 'N/A')}")
            with col2:
                if st.button("✅", key=f"completar_vencida_{idx}", help="Marcar como completada"):
                    # Eliminar la acción
                    df_int_actualizado = df_int.drop(row_idx)
                    if utils.escribir_excel(config.ARCHIVO_CRM, "INTERACCIONES", df_int_actualizado):
                        st.success("✅ Acción completada y eliminada")
                        st.cache_data.clear()
                        time.sleep(0.3)
                        st.rerun()
                    else:
                        st.error("❌ Error al actualizar")
    else:
        st.success("✅ No hay acciones vencidas")
    
    st.markdown("### 🟡 HOY")
    hoy_acc = df_acciones[df_acciones['Urgencia'] == 2]
    if not hoy_acc.empty:
        for idx, (row_idx, row) in enumerate(hoy_acc.iterrows()):
            col1, col2 = st.columns([9, 1])
            with col1:
                st.warning(f"📅 **HOY** - {row.get('Nombre Cliente', 'N/A')} - {row.get('Próxima Acción', 'N/A')}")
            with col2:
                if st.button("✅", key=f"completar_hoy_{idx}", help="Marcar como completada"):
                    df_int_actualizado = df_int.drop(row_idx)
                    if utils.escribir_excel(config.ARCHIVO_CRM, "INTERACCIONES", df_int_actualizado):
                        st.success("✅ Acción completada y eliminada")
                        st.cache_data.clear()
                        time.sleep(0.3)
                        st.rerun()
                    else:
                        st.error("❌ Error al actualizar")
    else:
        st.info("✅ No hay acciones para hoy")
    
    st.markdown("### 🟢 ESTA SEMANA")
    semana = df_acciones[df_acciones['Urgencia'] == 3]
    if not semana.empty:
        for idx, (row_idx, row) in enumerate(semana.iterrows()):
            try:
                fecha = pd.to_datetime(row.get('Fecha Próxima Acción')).strftime('%d/%m/%Y')
            except:
                fecha = str(row.get('Fecha Próxima Acción', 'N/A'))
            
            col1, col2 = st.columns([9, 1])
            with col1:
                st.info(f"📆 **{fecha}** - {row.get('Nombre Cliente', 'N/A')} - {row.get('Próxima Acción', 'N/A')}")
            with col2:
                if st.button("✅", key=f"completar_semana_{idx}", help="Marcar como completada"):
                    df_int_actualizado = df_int.drop(row_idx)
                    if utils.escribir_excel(config.ARCHIVO_CRM, "INTERACCIONES", df_int_actualizado):
                        st.success("✅ Acción completada y eliminada")
                        st.cache_data.clear()
                        time.sleep(0.3)
                        st.rerun()
                    else:
                        st.error("❌ Error al actualizar")
    else:
        st.info("✅ No hay acciones esta semana")
    
    st.markdown("### 🔵 PRÓXIMAMENTE")
    futuras = df_acciones[df_acciones['Urgencia'] == 4]
    if not futuras.empty:
        for idx, (row_idx, row) in enumerate(futuras.iterrows()):
            try:
                fecha = pd.to_datetime(row.get('Fecha Próxima Acción')).strftime('%d/%m/%Y')
            except:
                fecha = str(row.get('Fecha Próxima Acción', 'N/A'))
            
            col1, col2 = st.columns([9, 1])
            with col1:
                st.info(f"📅 **{fecha}** - {row.get('Nombre Cliente', 'N/A')} - {row.get('Próxima Acción', 'N/A')}")
            with col2:
                if st.button("✅", key=f"completar_futura_{idx}", help="Marcar como completada"):
                    df_int_actualizado = df_int.drop(row_idx)
                    if utils.escribir_excel(config.ARCHIVO_CRM, "INTERACCIONES", df_int_actualizado):
                        st.success("✅ Acción completada y eliminada")
                        st.cache_data.clear()
                        time.sleep(0.3)
                        st.rerun()
                    else:
                        st.error("❌ Error al actualizar")
    else:
        st.info("✅ No hay acciones programadas después de una semana")

# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================

def crear_interaccion_modal(id_cliente, nombre_cliente):
    """Modal para crear nueva interacción"""
    st.markdown("### ➕ Nueva Interacción")
    
    with st.form(key=f"form_interaccion_{id_cliente}"):
        fecha = st.date_input("Fecha:", value=datetime.now())
        tipo = st.selectbox("Tipo:", ['Llamada', 'Email', 'Reunión', 'WhatsApp', 'Visita', 'Otro'])
        descripcion = st.text_area("Descripción:", height=100)
        
        col1, col2 = st.columns(2)
        with col1:
            proxima_accion = st.text_input("Próxima Acción:")
        with col2:
            fecha_proxima = st.date_input("Fecha Próxima Acción:", value=datetime.now() + timedelta(days=7))
        
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            cancelar = st.form_submit_button("❌ Cancelar", use_container_width=True)
        
        with col_btn2:
            guardar = st.form_submit_button("💾 Guardar", use_container_width=True, type="primary")
        
        if cancelar:
            st.session_state['crear_interaccion'] = False
            st.rerun()
        
        if guardar:
            df_int = utils.leer_excel_forzado(config.ARCHIVO_CRM, "INTERACCIONES")
            
            nueva_interaccion = {
                'ID': utils.generar_id(),
                'ID Cliente': id_cliente,
                'Nombre Cliente': nombre_cliente,
                'Fecha': fecha,
                'Tipo': tipo,
                'Descripción': descripcion,
                'Próxima Acción': proxima_accion,
                'Fecha Próxima Acción': fecha_proxima if proxima_accion else None
            }
            
            df_int = pd.concat([df_int, pd.DataFrame([nueva_interaccion])], ignore_index=True)
            
            if utils.escribir_excel(config.ARCHIVO_CRM, "INTERACCIONES", df_int):
                st.success("✅ Interacción registrada")
                st.session_state['crear_interaccion'] = False
                st.cache_data.clear()
                time.sleep(0.5)
                st.rerun()
            else:
                st.error("❌ Error al guardar")

def crear_modal_whatsapp(id_cliente, nombre_cliente):
    """Modal para enviar WhatsApp personalizado al contacto principal"""
    st.markdown("### 📱 Enviar Mensaje WhatsApp")
    
    # Buscar contacto principal
    df_contactos = utils.leer_excel_forzado(config.ARCHIVO_CRM, "CONTACTOS")
    df_clientes = utils.leer_excel(config.ARCHIVO_CRM, "CLIENTES_ACTIVOS")
    cliente_info = df_clientes[df_clientes['ID'] == id_cliente]
    
    contacto_principal = None
    nombre_contacto = nombre_cliente
    numero_telefono = ""
    
    # Buscar contacto principal en CONTACTOS
    if not df_contactos.empty:
        contactos_cliente = df_contactos[df_contactos['ID Cliente'] == id_cliente]
        if not contactos_cliente.empty:
            # Convertir 'Es Principal' a bool si es necesario
            contactos_cliente = contactos_cliente.copy()
            contactos_cliente['Es Principal'] = contactos_cliente['Es Principal'].astype(bool)
            principal = contactos_cliente[contactos_cliente['Es Principal'] == True]
            if not principal.empty:
                contacto_principal = principal.iloc[0]
                nombre_contacto = contacto_principal.get('Nombre Contacto', nombre_cliente)
                numero_telefono = str(contacto_principal.get('Teléfono', ''))
    
    # Si no hay contacto principal, usar el teléfono del cliente
    if not numero_telefono:
        if not cliente_info.empty:
            numero_telefono = cliente_info.iloc[0].get('Teléfono', '')
        if not numero_telefono:
            st.error("❌ Este cliente no tiene teléfono registrado")
            if st.button("❌ Cerrar", use_container_width=True):
                st.session_state['enviar_whatsapp_modal'] = False
                st.rerun()
            return
    
    with st.form(key=f"form_whatsapp_{id_cliente}"):
        st.write(f"**Cliente:** {nombre_cliente}")
        st.write(f"**Contacto Principal:** {nombre_contacto}")
        st.write(f"**Teléfono:** {numero_telefono}")
        
        mensaje = st.text_area(
            "Mensaje:",
            height=200,
            placeholder="Escribe tu mensaje aquí...",
            value=f"Hola {nombre_contacto}, me gustaría comentarte sobre tu negocio..."
        )
        
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            cancelar = st.form_submit_button("❌ Cancelar", use_container_width=True)
        
        with col_btn2:
            enviar = st.form_submit_button("📱 Enviar WhatsApp", use_container_width=True, type="primary")
        
        if cancelar:
            st.session_state['enviar_whatsapp_modal'] = False
            st.rerun()
        
        if enviar:
            if not mensaje:
                st.error("❌ El mensaje no puede estar vacío")
            else:
                import urllib.parse
                import webbrowser
                
                # Limpiar número
                numero_limpio = numero_telefono.replace(" ", "").replace("-", "").replace("+", "")
                if not numero_limpio.startswith("34"):
                    numero_limpio = "34" + numero_limpio
                numero_final = "+" + numero_limpio
                
                # Codificar mensaje para URL
                mensaje_encoded = urllib.parse.quote(mensaje)
                url_whatsapp = f"https://web.whatsapp.com/send?phone={numero_final}&text={mensaje_encoded}"
                
                # Registrar en Timeline
                df_int = utils.leer_excel_forzado(config.ARCHIVO_CRM, "INTERACCIONES")
                
                nueva_interaccion = {
                    'ID': utils.generar_id(),
                    'ID Cliente': id_cliente,
                    'Nombre Cliente': nombre_cliente,
                    'Fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'Tipo': 'WhatsApp',
                    'Descripción': f"Mensaje a {nombre_contacto}: {mensaje[:100]}..." if len(mensaje) > 100 else f"Mensaje a {nombre_contacto}: {mensaje}",
                    'Próxima Acción': '',
                    'Fecha Próxima Acción': None
                }
                
                df_int = pd.concat([df_int, pd.DataFrame([nueva_interaccion])], ignore_index=True)
                utils.escribir_excel(config.ARCHIVO_CRM, "INTERACCIONES", df_int)
                
                # Abrir WhatsApp Web
                webbrowser.open(url_whatsapp)
                
                st.success(f"✅ WhatsApp abierto - El mensaje a {nombre_contacto} está listo, solo presiona Enter")
                st.session_state['enviar_whatsapp_modal'] = False
                st.cache_data.clear()
                st.rerun()

def programar_tarea_modal(id_cliente, nombre_cliente):
    """Modal para programar tarea - NUEVA FUNCIONALIDAD"""
    st.markdown("### 📅 Programar Nueva Tarea")
    
    with st.form(key=f"form_tarea_{id_cliente}"):
        st.info("Esta tarea se guardará como una próxima acción y aparecerá en la pestaña 'Próximas Acciones'")
        
        descripcion_tarea = st.text_area("Descripción de la Tarea:", height=100, 
                                        placeholder="Ej: Llamar para seguimiento del servicio")
        
        col1, col2 = st.columns(2)
        with col1:
            fecha_tarea = st.date_input("Fecha de la Tarea:", value=datetime.now() + timedelta(days=1))
        
        with col2:
            tipo_tarea = st.selectbox("Tipo:", ['Llamada', 'Email', 'Reunión', 'WhatsApp', 'Visita', 'Otro'])
        
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            cancelar = st.form_submit_button("❌ Cancelar", use_container_width=True)
        
        with col_btn2:
            guardar = st.form_submit_button("💾 Programar", use_container_width=True, type="primary")
        
        if cancelar:
            st.session_state['programar_tarea'] = False
            st.rerun()
        
        if guardar:
            if not descripcion_tarea:
                st.error("❌ La descripción de la tarea es obligatoria")
            else:
                df_int = utils.leer_excel_forzado(config.ARCHIVO_CRM, "INTERACCIONES")
                
                nueva_tarea = {
                    'ID': utils.generar_id(),
                    'ID Cliente': id_cliente,
                    'Nombre Cliente': nombre_cliente,
                    'Fecha': datetime.now(),
                    'Tipo': tipo_tarea,
                    'Descripción': f"TAREA PROGRAMADA: {descripcion_tarea}",
                    'Próxima Acción': descripcion_tarea,
                    'Fecha Próxima Acción': fecha_tarea
                }
                
                df_int = pd.concat([df_int, pd.DataFrame([nueva_tarea])], ignore_index=True)
                
                if utils.escribir_excel(config.ARCHIVO_CRM, "INTERACCIONES", df_int):
                    st.success("✅ Tarea programada correctamente")
                    st.session_state['programar_tarea'] = False
                    st.session_state['tarea_programada'] = True
                    st.cache_data.clear()
                    time.sleep(0.5)
                    st.rerun()
                else:
                    st.error("❌ Error al programar tarea")

def crear_contacto_modal(id_cliente, nombre_cliente):
    """Modal para crear nuevo contacto"""
    st.markdown("### ➕ Nuevo Contacto")
    
    with st.form(key=f"form_contacto_{id_cliente}"):
        col1, col2 = st.columns(2)
        
        with col1:
            nombre_contacto = st.text_input("Nombre:*")
            cargo = st.selectbox("Cargo:", ['Dueño/Gerente', 'Cocinero/Chef', 'Administrativo', 
                                           'Responsable Compras', 'Camarero', 'Otro'])
            telefono = st.text_input("Teléfono:")
        
        with col2:
            email = st.text_input("Email/Correo:")
            empresa = st.text_input("Empresa/Institución:")
            es_principal = st.checkbox("Contacto Principal")
        
        notas = st.text_area("Notas:", height=80)
        
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            cancelar = st.form_submit_button("❌ Cancelar", use_container_width=True)
        
        with col_btn2:
            guardar = st.form_submit_button("💾 Guardar", use_container_width=True, type="primary")
        
        if cancelar:
            st.session_state['crear_contacto'] = False
            st.rerun()
        
        if guardar:
            if not nombre_contacto:
                st.error("❌ El nombre del contacto es obligatorio")
            else:
                with st.spinner("Guardando contacto..."):
                    try:
                        df_contactos = utils.leer_excel_forzado(config.ARCHIVO_CRM, "CONTACTOS")
                    except:
                        df_contactos = pd.DataFrame()
                    
                    # Si no existe la hoja, crearla
                    if df_contactos.empty:
                        df_contactos = pd.DataFrame(columns=[
                            'ID', 'ID Cliente', 'Nombre Cliente', 'Nombre Contacto', 
                            'Cargo', 'Teléfono', 'Email', 'Empresa',
                            'Es Principal', 'Notas', 'Fecha Registro'
                        ])
                    
                    nuevo_contacto = {
                        'ID': utils.generar_id(),
                        'ID Cliente': id_cliente,
                        'Nombre Cliente': nombre_cliente,
                        'Nombre Contacto': nombre_contacto,
                        'Cargo': cargo,
                        'Teléfono': telefono,
                        'Email': email,
                        'Empresa': empresa,
                        'Es Principal': es_principal,
                        'Notas': notas,
                        'Fecha Registro': datetime.now()
                    }
                    
                    df_contactos = pd.concat([df_contactos, pd.DataFrame([nuevo_contacto])], ignore_index=True)
                    
                    # Limpiar caché ANTES de escribir
                    st.cache_data.clear()
                    
                    if utils.escribir_excel(config.ARCHIVO_CRM, "CONTACTOS", df_contactos):
                        time.sleep(1)
                        st.success("✅ Contacto añadido correctamente")
                        st.session_state['crear_contacto'] = False
                        st.cache_data.clear()
                        time.sleep(0.5)
                        st.rerun()
                    else:
                        st.error("❌ Error al guardar contacto")

def editar_contacto_modal(contacto, id_cliente):
    """Modal para editar contacto existente"""
    st.markdown("### ✏️ Editar Contacto")
    
    with st.form(key=f"form_editar_contacto_{contacto.get('ID', '')}"):
        col1, col2 = st.columns(2)
        
        with col1:
            nombre_contacto = st.text_input("Nombre:*", value=contacto.get('Nombre Contacto', ''))
            cargo = st.selectbox("Cargo:", ['Dueño/Gerente', 'Cocinero/Chef', 'Administrativo', 
                                           'Responsable Compras', 'Camarero', 'Otro'],
                                index=['Dueño/Gerente', 'Cocinero/Chef', 'Administrativo', 
                                      'Responsable Compras', 'Camarero', 'Otro'].index(contacto.get('Cargo', 'Dueño/Gerente')))
            telefono = st.text_input("Teléfono:", value=contacto.get('Teléfono', ''))
        
        with col2:
            email = st.text_input("Email/Correo:", value=contacto.get('Email', ''))
            empresa = st.text_input("Empresa/Institución:", value=contacto.get('Empresa', ''))
            es_principal = st.checkbox("Contacto Principal", value=bool(contacto.get('Es Principal', False)))
        
        notas = st.text_area("Notas:", value=contacto.get('Notas', ''), height=80)
        
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            cancelar = st.form_submit_button("❌ Cancelar", use_container_width=True)
        
        with col_btn2:
            guardar = st.form_submit_button("💾 Guardar", use_container_width=True, type="primary")
        
        if cancelar:
            st.session_state['editar_contacto'] = False
            st.rerun()
        
        if guardar:
            if not nombre_contacto:
                st.error("❌ El nombre del contacto es obligatorio")
            else:
                with st.spinner("Guardando cambios..."):
                    try:
                        df_contactos = utils.leer_excel_forzado(config.ARCHIVO_CRM, "CONTACTOS")
                    except:
                        df_contactos = pd.DataFrame()
                    
                    # Actualizar el contacto
                    contacto_id = contacto.get('ID')
                    mask = df_contactos['ID'] == contacto_id
                    
                    if mask.any():
                        df_contactos.loc[mask, 'Nombre Contacto'] = nombre_contacto
                        df_contactos.loc[mask, 'Cargo'] = cargo
                        df_contactos.loc[mask, 'Teléfono'] = telefono
                        df_contactos.loc[mask, 'Email'] = email
                        df_contactos.loc[mask, 'Empresa'] = empresa
                        df_contactos.loc[mask, 'Es Principal'] = es_principal
                        df_contactos.loc[mask, 'Notas'] = notas
                        
                        # Limpiar caché ANTES de escribir
                        st.cache_data.clear()
                        
                        if utils.escribir_excel(config.ARCHIVO_CRM, "CONTACTOS", df_contactos):
                            time.sleep(1)
                            st.success("✅ Contacto actualizado correctamente")
                            st.session_state['editar_contacto'] = False
                            st.cache_data.clear()
                            time.sleep(0.5)
                            st.rerun()
                        else:
                            st.error("❌ Error al guardar cambios")
                    else:
                        st.error("❌ Contacto no encontrado")

def crear_cliente_nuevo():
    """Modal para crear nuevo cliente"""
    st.markdown("### ➕ Nuevo Cliente")
    
    with st.form(key="form_nuevo_cliente"):
        col1, col2 = st.columns(2)
        
        with col1:
            nombre = st.text_input("Nombre Comercial:*")
            cif = st.text_input("CIF:")
            telefono = st.text_input("Teléfono:*")
            email = st.text_input("Email:")
        
        with col2:
            direccion = st.text_area("Dirección:", height=100)
            contacto = st.text_input("Persona de Contacto:")
            comercial = st.text_input("Comercial Asignado:")
            estado = st.selectbox("Estado:", ['Lead', 'Cliente Activo'])
        
        notas = st.text_area("Notas:", height=100)
        
        col_btn1, col_btn2, col_btn3 = st.columns([2, 1, 1])
        
        with col_btn2:
            cancelar = st.form_submit_button("Cancelar", use_container_width=True)
        
        with col_btn3:
            crear = st.form_submit_button("💾 Crear", use_container_width=True, type="primary")
        
        if cancelar:
            st.session_state['crear_nuevo_cliente'] = False
            st.rerun()
        
        if crear:
            if not nombre or not telefono:
                st.error("❌ Nombre y Teléfono son obligatorios")
            else:
                df_clientes = utils.leer_excel_forzado(config.ARCHIVO_CRM, "LEADS")
                
                nuevo_cliente = {
                    'ID': utils.generar_id(),
                    'Nombre Comercial': nombre,
                    'CIF': cif,
                    'Teléfono': telefono,
                    'Email': email,
                    'Dirección': direccion,
                    'Contacto': contacto,
                    'Comercial Asignado': comercial,
                    'Estado': estado,
                    'Fecha Alta': datetime.now(),
                    'Notas': notas,
                    'Fecha_Cambio_Estado': datetime.now(),
                    'Usuario': 'Sistema'
                }
                
                df_clientes = pd.concat([df_clientes, pd.DataFrame([nuevo_cliente])], ignore_index=True)
                
                st.cache_data.clear()
                if utils.escribir_excel(config.ARCHIVO_CRM, "LEADS", df_clientes):
                    st.success(f"✅ Cliente creado: {nombre}")
                    st.session_state['crear_nuevo_cliente'] = False
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error("❌ Error al crear cliente")

# ============================================================================
# FUNCIÓN MODAL: EDITAR CLIENTE INCOMPLETO
# ============================================================================

@st.dialog("✏️ Editar Cliente", width="large")
def editar_cliente_incompleto_modal():
    """Modal para editar rápidamente cliente incompleto (Dirección + Ciudad)"""
    
    id_cliente = st.session_state.get('id_cliente_editar')
    nombre_cliente = st.session_state.get('nombre_cliente_editar')
    
    if not id_cliente:
        st.error("❌ No se especificó cliente")
        return
    
    # Cargar datos del cliente
    df_clientes = utils.leer_excel_forzado(config.ARCHIVO_CRM, "CLIENTES_ACTIVOS")
    cliente_data = df_clientes[df_clientes['ID'] == id_cliente]
    
    if cliente_data.empty:
        st.error(f"❌ Cliente no encontrado (ID: {id_cliente})")
        return
    
    cliente = cliente_data.iloc[0]
    
    st.success(f"Editando: **{nombre_cliente}**")
    
    with st.form(key=f"form_editar_incompleto_{id_cliente}"):
        col1, col2 = st.columns(2)
        
        with col1:
            direccion = st.text_area(
                "📍 Dirección *",
                value=cliente.get('Dirección', ''),
                placeholder="C/ Principal, 123",
                help="Dirección para geocodificación automática"
            )
        
        with col2:
            ciudad = st.text_input(
                "🏙️ Ciudad *",
                value=cliente.get('Ciudad', ''),
                placeholder="Madrid, Barcelona...",
                help="Ciudad para geocodificación correcta"
            )
        
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            cancelar = st.form_submit_button("❌ Cancelar", use_container_width=True)
        
        with col_btn2:
            guardar = st.form_submit_button("💾 Guardar", use_container_width=True, type="primary")
        
        if cancelar:
            st.session_state['editar_cliente'] = False
            st.session_state.pop('id_cliente_editar', None)
            st.session_state.pop('nombre_cliente_editar', None)
            st.rerun()
        
        if guardar:
            # Validación
            errores = []
            if not direccion or not direccion.strip():
                errores.append("📍 **Dirección** es obligatoria")
            if not ciudad or not ciudad.strip():
                errores.append("🏙️ **Ciudad** es obligatoria")
            
            if errores:
                st.error("❌ Faltan campos:")
                for error in errores:
                    st.error(error)
            else:
                with st.spinner("Geocodificando y guardando..."):
                    # Geocodificar automáticamente
                    lat, lon = None, None
                    
                    try:
                        from geopy.geocoders import Nominatim
                        geolocator = Nominatim(user_agent="horeca_crm", timeout=5)
                        
                        direccion_completa = f"{direccion.strip()}, {ciudad.strip()}, España"
                        
                        try:
                            location = geolocator.geocode(direccion_completa, language='es')
                            if location:
                                lat = location.latitude
                                lon = location.longitude
                            else:
                                # Fallback a caché
                                ciudades_cache = {
                                    'madrid': (40.4168, -3.7038),
                                    'barcelona': (41.3874, 2.1686),
                                    'valencia': (39.4699, -0.3763),
                                }
                                ciudad_lower = ciudad.lower().strip()
                                lat, lon = ciudades_cache.get(ciudad_lower, (40.4168, -3.7038))
                        except Exception:
                            ciudades_cache = {
                                'madrid': (40.4168, -3.7038),
                                'barcelona': (41.3874, 2.1686),
                                'valencia': (39.4699, -0.3763),
                            }
                            ciudad_lower = ciudad.lower().strip()
                            lat, lon = ciudades_cache.get(ciudad_lower, (40.4168, -3.7038))
                    except ImportError:
                        ciudades_cache = {
                            'madrid': (40.4168, -3.7038),
                            'barcelona': (41.3874, 2.1686),
                            'valencia': (39.4699, -0.3763),
                        }
                        ciudad_lower = ciudad.lower().strip()
                        lat, lon = ciudades_cache.get(ciudad_lower, (40.4168, -3.7038))
                    
                    # Actualizar cliente
                    mascara = df_clientes['ID'] == id_cliente
                    df_clientes.loc[mascara, 'Dirección'] = direccion
                    df_clientes.loc[mascara, 'Ciudad'] = ciudad
                    df_clientes.loc[mascara, 'Latitud'] = float(lat) if lat else None
                    df_clientes.loc[mascara, 'Longitud'] = float(lon) if lon else None
                    
                    # Guardar
                    st.cache_data.clear()
                    if utils.escribir_excel(config.ARCHIVO_CRM, "CLIENTES_ACTIVOS", df_clientes):
                        st.success("✅ Cliente actualizado correctamente")
                        st.success(f"📍 Geocodificado en: ({lat:.4f}, {lon:.4f})")
                        time.sleep(0.5)
                        st.session_state['editar_cliente'] = False
                        st.session_state.pop('id_cliente_editar', None)
                        st.session_state.pop('nombre_cliente_editar', None)
                        st.rerun()
                    else:
                        st.error("❌ Error al guardar")

# ============================================================================
# FUNCIÓN MODAL: CONVERTIR LEAD A CLIENTE ACTIVO
# ============================================================================
# Esta función está implementada en funciones_leads.py
# Se importa automáticamente al inicio del módulo
# No se debe duplicar aquí
