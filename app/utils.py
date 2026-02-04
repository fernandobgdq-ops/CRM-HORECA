"""
UTILS.PY - Funciones Utilitarias
Lectura/Escritura de Excel y funciones comunes
"""

import pandas as pd
import streamlit as st
from datetime import datetime, date
import config
from io import BytesIO

# ============================================================================
# FUNCIONES DE LECTURA DE EXCEL
# ============================================================================

def leer_excel(archivo, hoja):
    """
    Lee una hoja de Excel y la devuelve como DataFrame
    
    Args:
        archivo: Ruta del archivo Excel
        hoja: Nombre de la hoja a leer
    
    Returns:
        DataFrame con los datos
    """
    try:
        if config.USE_ONEDRIVE_API:
            import onedrive
            contenido = onedrive.descargar_archivo(archivo)
            df = pd.read_excel(BytesIO(contenido), sheet_name=hoja)
        else:
            df = pd.read_excel(archivo, sheet_name=hoja)
        return df
    except Exception:
        # Retornar DataFrame vacío sin mostrar error (para hojas opcionales)
        return pd.DataFrame()

def leer_todas_hojas(archivo):
    """Lee todas las hojas de un archivo Excel"""
    try:
        if config.USE_ONEDRIVE_API:
            import onedrive
            contenido = onedrive.descargar_archivo(archivo)
            excel_file = pd.ExcelFile(BytesIO(contenido))
        else:
            excel_file = pd.ExcelFile(archivo)
        hojas = {}
        for nombre_hoja in excel_file.sheet_names:
            if config.USE_ONEDRIVE_API:
                hojas[nombre_hoja] = pd.read_excel(BytesIO(contenido), sheet_name=nombre_hoja)
            else:
                hojas[nombre_hoja] = pd.read_excel(archivo, sheet_name=nombre_hoja)
        return hojas
    except Exception as e:
        st.error(f"Error al leer {archivo}: {str(e)}")
        return {}

def leer_excel_forzado(archivo, hoja):
    """
    Lee una hoja de Excel forzando limpieza de caché sin usar pandas.cache
    Usa openpyxl directamente para evitar caché de Streamlit
    
    Args:
        archivo: Ruta del archivo Excel
        hoja: Nombre de la hoja a leer
    
    Returns:
        DataFrame con los datos frescos del disco
    """
    from openpyxl import load_workbook
    import time
    
    st.cache_data.clear()
    time.sleep(0.3)  # Pequeña pausa para asegurar limpieza
    
    try:
        if config.USE_ONEDRIVE_API:
            import onedrive
            contenido = onedrive.descargar_archivo(archivo)
            wb = load_workbook(BytesIO(contenido), data_only=True)
        else:
            # Cargar workbook con openpyxl (no usa caché de Streamlit)
            wb = load_workbook(archivo, data_only=True)
        
        if hoja not in wb.sheetnames:
            st.error(f"Hoja '{hoja}' no encontrada en {archivo}")
            return pd.DataFrame()
        
        ws = wb[hoja]
        
        # Extraer datos
        datos = []
        encabezados = None
        
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                encabezados = row
            else:
                datos.append(row)
        
        # Crear DataFrame
        if encabezados:
            df = pd.DataFrame(datos, columns=encabezados)
        else:
            df = pd.DataFrame()
        
        wb.close()
        return df
        
    except Exception as e:
        st.error(f"Error al leer {archivo} - {hoja}: {str(e)}")
        return pd.DataFrame()

# ============================================================================
# FUNCIONES DE ESCRITURA EN EXCEL
# ============================================================================

def escribir_excel(archivo, hoja, df):
    """
    Escribe un DataFrame en una hoja específica de Excel
    Preserva las otras hojas del archivo
    IMPORTANTE: Llama st.cache_data.clear() ANTES de esta función
    
    Args:
        archivo: Ruta del archivo Excel
        hoja: Nombre de la hoja a escribir
        df: DataFrame a escribir
    
    Returns:
        True si fue exitoso, False si no
    """
    import time
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    max_intentos = 5
    intento = 0
    
    while intento < max_intentos:
        try:
            # Opción 1: Usar openpyxl directamente para mayor control
            # Cargar el libro existente
            if config.USE_ONEDRIVE_API:
                import onedrive
                try:
                    contenido = onedrive.descargar_archivo(archivo)
                    libro = openpyxl.load_workbook(BytesIO(contenido))
                except FileNotFoundError:
                    libro = openpyxl.Workbook()
                    if 'Sheet' in libro.sheetnames:
                        del libro['Sheet']
            else:
                try:
                    libro = openpyxl.load_workbook(archivo)
                except:
                    # Si no existe, crear uno nuevo
                    libro = openpyxl.Workbook()
                    if 'Sheet' in libro.sheetnames:
                        del libro['Sheet']
            
            # Eliminar la hoja si ya existe
            if hoja in libro.sheetnames:
                del libro[hoja]
            
            # Crear nueva hoja
            ws = libro.create_sheet(hoja)
            
            # Escribir headers
            for col_num, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_num, value=col_name)
            
            # Escribir datos
            for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar el libro
            if config.USE_ONEDRIVE_API:
                buffer = BytesIO()
                libro.save(buffer)
                libro.close()
                onedrive.subir_archivo(archivo, buffer.getvalue())
            else:
                libro.save(archivo)
                libro.close()
            
            # Verificar que se escribió correctamente
            time.sleep(0.5)
            if config.USE_ONEDRIVE_API:
                import onedrive
                contenido_verif = onedrive.descargar_archivo(archivo)
                df_verificacion = pd.read_excel(BytesIO(contenido_verif), sheet_name=hoja)
            else:
                df_verificacion = pd.read_excel(archivo, sheet_name=hoja)
            
            if len(df_verificacion) == len(df):
                print(f"[DEBUG] ✅ Excel guardado: {hoja} ({len(df)} filas)")
                return True
            else:
                print(f"[DEBUG] ⚠️ Verificación fallida: esperadas {len(df)} filas, se guardaron {len(df_verificacion)}")
                intento += 1
                if intento < max_intentos:
                    time.sleep(1)
                    continue
                else:
                    return False
                    
        except PermissionError as pe:
            intento += 1
            if intento < max_intentos:
                print(f"[DEBUG] 🔒 Archivo bloqueado, reintentando ({intento}/{max_intentos})...")
                time.sleep(2)
            else:
                st.error(f"❌ El archivo está bloqueado. Cierra Excel y espera a que OneDrive sincronice.")
                return False
        except Exception as e:
            intento += 1
            print(f"[DEBUG] Error (intento {intento}/{max_intentos}): {str(e)}")
            if intento < max_intentos:
                time.sleep(1)
            else:
                st.error(f"Error al escribir en Excel: {str(e)}")
                import traceback
                print(traceback.format_exc())
                return False
    
    return False

def agregar_fila(archivo, hoja, nueva_fila):
    """
    Agrega una nueva fila a una hoja de Excel
    
    Args:
        archivo: Ruta del archivo
        hoja: Nombre de la hoja
        nueva_fila: Dict con los datos de la nueva fila
    """
    try:
        # Leer la hoja actual directamente
        df = pd.read_excel(archivo, sheet_name=hoja)
        
        print(f"[DEBUG] Agregando fila a {hoja}")
        print(f"[DEBUG] Nombre: {nueva_fila.get('Nombre Comercial', nueva_fila.get('Nombre', 'N/A'))}")
        print(f"[DEBUG] Filas antes: {len(df)}")
        
        # Agregar la nueva fila
        nuevo_df = pd.concat([df, pd.DataFrame([nueva_fila])], ignore_index=True)
        
        print(f"[DEBUG] Filas después: {len(nuevo_df)}")
        
        # Escribir de vuelta
        resultado = escribir_excel(archivo, hoja, nuevo_df)
        
        if resultado:
            print(f"[DEBUG] ✅ Fila agregada y guardada en {hoja}")
            
            # Verificar que se guardó
            df_verif = pd.read_excel(archivo, sheet_name=hoja)
            print(f"[DEBUG] Verificación: ahora hay {len(df_verif)} filas en {hoja}")
        else:
            print(f"[DEBUG] ❌ Error al guardar en {hoja}")
            
        return resultado
        
    except Exception as e:
        print(f"[DEBUG] ❌ Excepción al agregar fila: {str(e)}")
        st.error(f"Error al agregar fila: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return False

def actualizar_fila(archivo, hoja, indice, columna, nuevo_valor):
    """
    Actualiza un valor específico en una fila
    
    Args:
        archivo: Ruta del archivo
        hoja: Nombre de la hoja
        indice: Índice de la fila (número de ID típicamente)
        columna: Nombre de la columna a actualizar
        nuevo_valor: Nuevo valor
    """
    try:
        df = leer_excel(archivo, hoja)
        df.loc[df.iloc[:, 0] == indice, columna] = nuevo_valor
        return escribir_excel(archivo, hoja, df)
    except Exception as e:
        st.error(f"Error al actualizar: {str(e)}")
        return False

def eliminar_fila(archivo, hoja, indice):
    """
    Elimina una fila basándose en el ID (primera columna)
    
    Args:
        archivo: Ruta del archivo
        hoja: Nombre de la hoja
        indice: ID de la fila a eliminar
    """
    try:
        df = leer_excel(archivo, hoja)
        df = df[df.iloc[:, 0] != indice]
        return escribir_excel(archivo, hoja, df)
    except Exception as e:
        st.error(f"Error al eliminar: {str(e)}")
        return False

# ============================================================================
# FUNCIONES DE CÁLCULO Y RETROALIMENTACIÓN
# ============================================================================

def actualizar_precio_mercado(id_ingrediente, nuevo_precio):
    """
    Actualiza el precio de mercado de un ingrediente
    y recalcula los escandallos afectados
    """
    try:
        # 1. Actualizar precio en INGREDIENTES_MAESTRO
        df_ing = leer_excel(config.ARCHIVO_OPERACIONES, "INGREDIENTES_MAESTRO")
        df_ing.loc[df_ing['ID Ingrediente'] == id_ingrediente, 'Precio Mercado Medio'] = nuevo_precio
        df_ing.loc[df_ing['ID Ingrediente'] == id_ingrediente, 'Última Actualización'] = datetime.now()
        escribir_excel(config.ARCHIVO_OPERACIONES, "INGREDIENTES_MAESTRO", df_ing)
        
        # 2. Actualizar escandallos que usan ese ingrediente
        df_esc = leer_excel(config.ARCHIVO_OPERACIONES, "ESCANDALLOS")
        mascara = df_esc['ID Ingrediente'] == id_ingrediente
        df_esc.loc[mascara, 'Coste Unitario'] = nuevo_precio
        df_esc.loc[mascara, 'Última Actualización'] = datetime.now()
        escribir_excel(config.ARCHIVO_OPERACIONES, "ESCANDALLOS", df_esc)
        
        # 3. Recalcular costes de platos afectados
        recalcular_costes_platos(df_esc)
        
        return True
    except Exception as e:
        st.error(f"Error al actualizar precio: {str(e)}")
        return False

def recalcular_costes_platos(df_escandallos):
    """
    Recalcula el coste total de todos los platos
    basándose en sus escandallos
    """
    try:
        # Agrupar por plato y sumar costes
        costes_por_plato = df_escandallos.groupby('ID Plato').agg({
            'Coste Total': 'sum'
        }).reset_index()
        
        # Actualizar en CARTA_CLIENTES
        df_carta = leer_excel(config.ARCHIVO_OPERACIONES, "CARTA_CLIENTES")
        
        for _, row in costes_por_plato.iterrows():
            id_plato = row['ID Plato']
            nuevo_coste = row['Coste Total']
            
            mascara = df_carta['ID Plato'] == id_plato
            df_carta.loc[mascara, 'Coste Total'] = nuevo_coste
            
            # Recalcular márgenes
            precio_venta = df_carta.loc[mascara, 'Precio Venta'].values[0]
            if precio_venta > 0:
                margen_euros = precio_venta - nuevo_coste
                margen_pct = (margen_euros / precio_venta) * 100
                food_cost = (nuevo_coste / precio_venta) * 100
                
                df_carta.loc[mascara, 'Margen €'] = margen_euros
                df_carta.loc[mascara, 'Margen %'] = margen_pct
                df_carta.loc[mascara, 'Food Cost %'] = food_cost
        
        escribir_excel(config.ARCHIVO_OPERACIONES, "CARTA_CLIENTES", df_carta)
        return True
    except Exception as e:
        st.error(f"Error al recalcular costes: {str(e)}")
        return False

def detectar_alertas_precios():
    """
    Detecta si hay ingredientes comprados muy por encima del precio de mercado
    Retorna lista de alertas
    """
    alertas = []
    
    try:
        df_lineas = leer_excel(config.ARCHIVO_OPERACIONES, "LINEAS_COMPRA")
        df_ingredientes = leer_excel(config.ARCHIVO_OPERACIONES, "INGREDIENTES_MAESTRO")
        
        for _, linea in df_lineas.iterrows():
            id_ing = linea['ID Ingrediente']
            precio_pagado = linea['Precio Unitario']
            
            # Buscar precio de mercado
            ing = df_ingredientes[df_ingredientes['ID Ingrediente'] == id_ing]
            if not ing.empty:
                precio_mercado = ing['Precio Mercado Medio'].values[0]
                
                if precio_mercado > 0:
                    desviacion = ((precio_pagado - precio_mercado) / precio_mercado) * 100
                    
                    if desviacion > config.UMBRAL_DESVIACION_PRECIO:
                        alertas.append({
                            'tipo': 'PRECIO_ALTO',
                            'ingrediente': linea['Nombre Ingrediente'],
                            'precio_pagado': precio_pagado,
                            'precio_mercado': precio_mercado,
                            'desviacion': round(desviacion, 1),
                            'ahorro_potencial': (precio_pagado - precio_mercado) * linea['Cantidad']
                        })
        
        return alertas
    except Exception as e:
        st.error(f"Error al detectar alertas: {str(e)}")
        return []

def detectar_alertas_margenes():
    """
    Detecta platos con márgenes peligrosamente bajos
    """
    alertas = []
    
    try:
        df_carta = leer_excel(config.ARCHIVO_OPERACIONES, "CARTA_CLIENTES")
        
        for _, plato in df_carta.iterrows():
            if plato['Activo'] == 'Sí':
                margen = plato['Margen %']
                
                if margen < config.UMBRAL_MARGEN_MINIMO:
                    alertas.append({
                        'tipo': 'MARGEN_BAJO',
                        'cliente': plato['Nombre Cliente'],
                        'plato': plato['Nombre Plato'],
                        'margen_actual': round(margen, 1),
                        'precio_venta': plato['Precio Venta'],
                        'coste': plato['Coste Total']
                    })
        
        return alertas
    except Exception as e:
        st.error(f"Error al detectar alertas de margen: {str(e)}")
        return []

# ============================================================================
# FUNCIONES DE FORMATEO
# ============================================================================

def formatear_moneda(valor):
    """Formatea un número como moneda"""
    try:
        return f"{float(valor):,.2f} €"
    except:
        return "0,00 €"

def formatear_porcentaje(valor):
    """Formatea un número como porcentaje"""
    try:
        return f"{float(valor):.1f}%"
    except:
        return "0.0%"

def color_margen(margen):
    """Devuelve color según el margen"""
    try:
        margen = float(margen)
        if margen < 20:
            return 'red'
        elif margen < 30:
            return 'orange'
        else:
            return 'green'
    except:
        return 'black'

# ============================================================================
# FUNCIONES DE ID AUTOMÁTICO
# ============================================================================

def obtener_siguiente_id(archivo, hoja):
    """
    Obtiene el siguiente ID disponible en una hoja
    Asume que la primera columna es el ID
    """
    try:
        df = leer_excel(archivo, hoja)
        if df.empty:
            return 1
        
        max_id = df.iloc[:, 0].max()
        return int(max_id) + 1 if pd.notna(max_id) else 1
    except:
        return 1

# ============================================================================
# FUNCIONES DE VALIDACIÓN
# ============================================================================

def validar_email(email):
    """Valida formato de email"""
    import re
    patron = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(patron, email) is not None

def validar_telefono(telefono):
    """Valida formato de teléfono español"""
    import re
    patron = r'^[6-9]\d{8}$'
    telefono_limpio = ''.join(filter(str.isdigit, str(telefono)))
    return re.match(patron, telefono_limpio) is not None

def validar_cif(cif):
    """Valida formato de CIF español (básico)"""
    import re
    patron = r'^[A-Z]\d{8}$'
    return re.match(patron, str(cif).upper()) is not None

# ============================================================================
# FUNCIONES DE FECHA
# ============================================================================

def fecha_a_texto(fecha):
    """Convierte fecha a texto legible"""
    if pd.isna(fecha):
        return ""
    try:
        if isinstance(fecha, str):
            return fecha
        return fecha.strftime("%d/%m/%Y")
    except:
        return ""

def texto_a_fecha(texto):
    """Convierte texto a fecha"""
    try:
        return datetime.strptime(texto, "%d/%m/%Y").date()
    except:
        return None

# ============================================================================
# FUNCIONES DE PROVEEDORES
# ============================================================================

@st.cache_data(ttl=5)
def obtener_nombre_proveedor(id_proveedor):
    """
    Obtiene el nombre comercial de un proveedor por su ID
    
    Args:
        id_proveedor: ID del proveedor
    
    Returns:
        Nombre comercial o "Desconocido" si no existe
    """
    try:
        if pd.isna(id_proveedor) or id_proveedor is None:
            return "Sin asignar"
        
        df_prov = leer_excel(config.ARCHIVO_PROVEEDORES, "PROVEEDORES")
        proveedor = df_prov[df_prov['ID Proveedor'] == int(id_proveedor)]
        
        if not proveedor.empty:
            return proveedor['Nombre Comercial'].values[0]
        else:
            return f"Proveedor ID {id_proveedor} (no encontrado)"
    except Exception as e:
        print(f"[DEBUG] Error obteniendo nombre proveedor: {e}")
        return "Error"

@st.cache_data(ttl=5)
def obtener_lista_proveedores():
    """
    Obtiene lista de todos los proveedores activos
    
    Returns:
        Lista de tuplas (ID, Nombre Comercial)
    """
    try:
        df_prov = leer_excel(config.ARCHIVO_PROVEEDORES, "PROVEEDORES")
        
        # Filtrar solo activos
        if 'Activo' in df_prov.columns:
            df_prov = df_prov[df_prov['Activo'] == 'Sí']
        
        # Crear lista de tuplas
        proveedores = [(row['ID Proveedor'], row['Nombre Comercial']) 
                       for _, row in df_prov.iterrows()]
        
        return proveedores
    except Exception as e:
        st.error(f"Error al obtener proveedores: {str(e)}")
        return []

def obtener_dict_proveedores():
    """
    Obtiene diccionario de proveedores {ID: Nombre}
    
    Returns:
        Dict con ID como clave y Nombre como valor
    """
    try:
        lista = obtener_lista_proveedores()
        return {id_prov: nombre for id_prov, nombre in lista}
    except:
        return {}

def generar_id():
    import uuid
    return str(uuid.uuid4())[:8]

# ============================================================================
# FUNCIONES DE INVENTARIO Y SNAPSHOTS
# ============================================================================

def obtener_inventario_cliente(id_cliente):
    """
    Obtiene el inventario actual de un cliente
    
    Args:
        id_cliente: ID del cliente
    
    Returns:
        DataFrame con el inventario del cliente
    """
    try:
        # Intentar leer la hoja INVENTARIO_CLIENTES
        try:
            df_inventario = leer_excel(config.ARCHIVO_OPERACIONES, "INVENTARIO_CLIENTES")
        except:
            # Si la hoja no existe, retornar DataFrame vacío
            return pd.DataFrame(columns=['ID Cliente', 'ID Ingrediente', 'Nombre Ingrediente', 
                                        'Cantidad Stock', 'Unidad', 'Última Actualización'])
        
        if df_inventario.empty:
            return pd.DataFrame(columns=['ID Cliente', 'ID Ingrediente', 'Nombre Ingrediente', 
                                        'Cantidad Stock', 'Unidad', 'Última Actualización'])
        
        # Filtrar por cliente
        inventario_cliente = df_inventario[df_inventario['ID Cliente'] == id_cliente].copy()
        
        return inventario_cliente
        
    except Exception as e:
        return pd.DataFrame()

def guardar_inventario_cliente(id_cliente, df_inventario):
    """
    Guarda o actualiza el inventario de un cliente
    
    Args:
        id_cliente: ID del cliente
        df_inventario: DataFrame con el inventario actualizado
    
    Returns:
        True si se guardó correctamente, False en caso contrario
    """
    try:
        # Leer todo el inventario (si existe la hoja)
        try:
            df_inventario_completo = leer_excel(config.ARCHIVO_OPERACIONES, "INVENTARIO_CLIENTES")
        except:
            # Si la hoja no existe, comenzar con un DataFrame vacío
            df_inventario_completo = pd.DataFrame(columns=['ID Cliente', 'ID Ingrediente', 'Nombre Ingrediente', 
                                                          'Cantidad Stock', 'Unidad', 'Última Actualización'])
        
        # Eliminar registros antiguos del cliente
        if not df_inventario_completo.empty:
            df_inventario_completo = df_inventario_completo[df_inventario_completo['ID Cliente'] != id_cliente]
        
        # Añadir el nuevo inventario
        df_inventario['ID Cliente'] = id_cliente
        df_inventario['Última Actualización'] = datetime.now()
        
        df_nuevo = pd.concat([df_inventario_completo, df_inventario], ignore_index=True)
        
        # Guardar
        return escribir_excel(config.ARCHIVO_OPERACIONES, "INVENTARIO_CLIENTES", df_nuevo)
        
    except Exception as e:
        return False

def guardar_snapshot(id_cliente, nombre_cliente, df_pedido, motivo=""):
    """
    Guarda un snapshot (instantánea) de un pedido
    
    Args:
        id_cliente: ID del cliente
        nombre_cliente: Nombre del cliente
        df_pedido: DataFrame con los productos del pedido
        motivo: Motivo del snapshot (opcional)
    
    Returns:
        True si se guardó correctamente
    """
    try:
        # Leer snapshots existentes
        try:
            df_snapshots = leer_excel(config.ARCHIVO_OPERACIONES, "SNAPSHOTS_PEDIDOS")
        except:
            df_snapshots = pd.DataFrame(columns=['ID Snapshot', 'ID Cliente', 'Nombre Cliente', 
                                                 'Fecha', 'Motivo', 'Total Productos', 
                                                 'Coste Total', 'Datos JSON'])
        
        # Crear nuevo snapshot
        snapshot_id = generar_id()
        fecha_actual = datetime.now()
        
        # Convertir pedido a JSON para guardar
        import json
        datos_json = df_pedido.to_json(orient='records')
        
        nuevo_snapshot = {
            'ID Snapshot': snapshot_id,
            'ID Cliente': id_cliente,
            'Nombre Cliente': nombre_cliente,
            'Fecha': fecha_actual,
            'Motivo': motivo,
            'Total Productos': len(df_pedido),
            'Coste Total': df_pedido['Coste Total'].sum() if 'Coste Total' in df_pedido.columns else 0,
            'Datos JSON': datos_json
        }
        
        # Agregar
        df_snapshots = pd.concat([df_snapshots, pd.DataFrame([nuevo_snapshot])], ignore_index=True)
        
        # Guardar
        return escribir_excel(config.ARCHIVO_OPERACIONES, "SNAPSHOTS_PEDIDOS", df_snapshots)
        
    except Exception as e:
        st.error(f"Error al guardar snapshot: {str(e)}")
        return False

def obtener_snapshots_cliente(id_cliente):
    """
    Obtiene todos los snapshots de un cliente
    
    Args:
        id_cliente: ID del cliente
    
    Returns:
        DataFrame con los snapshots del cliente
    """
    try:
        df_snapshots = leer_excel(config.ARCHIVO_OPERACIONES, "SNAPSHOTS_PEDIDOS")
        
        if df_snapshots.empty:
            return pd.DataFrame()
        
        # Filtrar por cliente
        snapshots_cliente = df_snapshots[df_snapshots['ID Cliente'] == id_cliente].copy()
        
        # Ordenar por fecha descendente
        if not snapshots_cliente.empty and 'Fecha' in snapshots_cliente.columns:
            snapshots_cliente = snapshots_cliente.sort_values('Fecha', ascending=False)
        
        return snapshots_cliente
        
    except Exception as e:
        st.error(f"Error al obtener snapshots: {str(e)}")
        return pd.DataFrame()